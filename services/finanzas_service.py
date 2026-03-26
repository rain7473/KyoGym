"""Servicio de finanzas: ingresos, egresos, morosidad y reportes"""
from datetime import date, timedelta
from pathlib import Path
from db import get_connection
from utils.constants import ESTADO_VENCIDA


CATEGORIAS_EGRESO = ["Alquiler", "Servicios", "Sueldos", "Mantenimiento", "Inventario", "Otro"]


def _resolver_reportes_dir() -> Path:
    """Resuelve el directorio de reportes buscando la unidad G: con
    los nombres posibles de Google Drive ('Mi unidad' en español o
    'My Drive' en inglés). Si ninguno existe, devuelve la carpeta
    local 'reportes'."""
    candidatos = [
        Path("G:/Mi unidad"),
        Path("G:/My Drive"),
    ]
    for candidato in candidatos:
        if candidato.exists():
            return candidato
    # Fallback: carpeta local junto al ejecutable
    return Path("reportes")


REPORTES_DIR = _resolver_reportes_dir()


# ─────────────────────────── INGRESOS ────────────────────────────

def listar_ingresos(fecha_desde=None, fecha_hasta=None, cliente=None, limite=2000):
    """Devuelve la lista de pagos (ingresos) con filtros opcionales."""
    conn = get_connection()
    cur = conn.cursor()

    query = """
        SELECT p.id, p.fecha, c.nombre AS cliente_nombre, c.telefono AS cliente_telefono,
               p.concepto, p.metodo, p.monto, p.cliente_id
        FROM pagos p
        JOIN clientes c ON p.cliente_id = c.id
        WHERE 1=1
    """
    params = []

    if fecha_desde:
        query += " AND p.fecha >= ?"
        params.append(fecha_desde if isinstance(fecha_desde, str) else fecha_desde.isoformat())
    if fecha_hasta:
        query += " AND p.fecha <= ?"
        params.append(fecha_hasta if isinstance(fecha_hasta, str) else fecha_hasta.isoformat())
    if cliente:
        query += " AND LOWER(c.nombre) LIKE ?"
        params.append(f"%{cliente.lower()}%")

    query += " ORDER BY p.fecha DESC, p.id DESC LIMIT ?"
    params.append(limite)

    cur.execute(query, params)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


def calcular_total_ingresos(fecha_desde=None, fecha_hasta=None):
    """Suma de ingresos en el período."""
    pagos = listar_ingresos(fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)
    return sum(p["monto"] for p in pagos)


# ─────────────────────────── EGRESOS ─────────────────────────────

def registrar_egreso(fecha, categoria, descripcion, proveedor, metodo, monto):
    """Registra un gasto. Devuelve el id del egreso creado."""
    if isinstance(fecha, date):
        fecha = fecha.isoformat()
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO egresos (fecha, categoria, descripcion, proveedor, metodo, monto)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (fecha, categoria, descripcion or "", proveedor or "", metodo, float(monto)))
    egreso_id = cur.lastrowid
    conn.commit()
    conn.close()
    return egreso_id


def listar_egresos(fecha_desde=None, fecha_hasta=None, categoria=None, limite=2000):
    """Devuelve la lista de egresos con filtros opcionales."""
    conn = get_connection()
    cur = conn.cursor()

    query = "SELECT * FROM egresos WHERE 1=1"
    params = []

    if fecha_desde:
        query += " AND fecha >= ?"
        params.append(fecha_desde if isinstance(fecha_desde, str) else fecha_desde.isoformat())
    if fecha_hasta:
        query += " AND fecha <= ?"
        params.append(fecha_hasta if isinstance(fecha_hasta, str) else fecha_hasta.isoformat())
    if categoria and categoria != "Todas":
        query += " AND categoria = ?"
        params.append(categoria)

    query += " ORDER BY fecha DESC, id DESC LIMIT ?"
    params.append(limite)

    cur.execute(query, params)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


def calcular_total_egresos(fecha_desde=None, fecha_hasta=None, categoria=None):
    """Suma de egresos en el período."""
    egresos = listar_egresos(fecha_desde=fecha_desde, fecha_hasta=fecha_hasta, categoria=categoria)
    return sum(e["monto"] for e in egresos)


def eliminar_egreso(egreso_id):
    """Elimina un egreso por ID."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM egresos WHERE id = ?", (egreso_id,))
    conn.commit()
    conn.close()


# ─────────────────────────── RESUMEN ─────────────────────────────

def obtener_resumen_mes(año=None, mes=None):
    """Devuelve un diccionario con las métricas del mes."""
    hoy = date.today()
    if año is None:
        año = hoy.year
    if mes is None:
        mes = hoy.month

    fecha_desde = date(año, mes, 1)
    if mes == 12:
        fecha_hasta = date(año, 12, 31)
    else:
        fecha_hasta = date(año, mes + 1, 1) - timedelta(days=1)

    ingresos_mes = calcular_total_ingresos(fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)
    egresos_mes = calcular_total_egresos(fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)
    utilidad = ingresos_mes - egresos_mes

    # Membresías activas y vencidas
    from services.membresia_service import listar_membresias, calcular_estado_membresia
    from utils.constants import ESTADO_ACTIVA, ESTADO_POR_VENCER

    todas = listar_membresias()
    activas = sum(1 for m in todas if m["estado"] in (ESTADO_ACTIVA, ESTADO_POR_VENCER))
    vencidas = sum(1 for m in todas if m["estado"] == ESTADO_VENCIDA)

    # Últimos 5 ingresos y egresos
    ultimos_ingresos = listar_ingresos(limite=5)
    ultimos_egresos = listar_egresos(limite=5)

    return {
        "ingresos_mes": ingresos_mes,
        "egresos_mes": egresos_mes,
        "utilidad": utilidad,
        "membresias_activas": activas,
        "cuentas_por_cobrar": vencidas,
        "ultimos_ingresos": ultimos_ingresos,
        "ultimos_egresos": ultimos_egresos,
    }


# ─────────────────────────── MOROSIDAD ───────────────────────────

def listar_morosos():
    """Devuelve clientes con membresía vencida, ordenados por días de atraso DESC."""
    from services.membresia_service import listar_membresias
    hoy = date.today()

    vencidas = []
    for m in listar_membresias(estado=ESTADO_VENCIDA):
        fv = date.fromisoformat(m["fecha_vencimiento"])
        dias = (hoy - fv).days
        m["dias_atraso"] = dias
        vencidas.append(m)

    vencidas.sort(key=lambda x: x["dias_atraso"], reverse=True)
    return vencidas


# ─────────────────────────── COMPARACIÓN MESES ───────────────────

def obtener_comparacion_meses(año=None):
    """Devuelve lista de dicts {mes, nombre_mes, ingresos, egresos, utilidad, variacion_pct}
    para cada mes del año."""
    if año is None:
        año = date.today().year

    nombres = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
               "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    resultado = []
    utilidad_anterior = None

    for mes in range(1, 13):
        fecha_desde = date(año, mes, 1)
        if mes == 12:
            fecha_hasta = date(año, 12, 31)
        else:
            fecha_hasta = date(año, mes + 1, 1) - timedelta(days=1)

        ingresos = calcular_total_ingresos(fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)
        egresos = calcular_total_egresos(fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)
        utilidad = ingresos - egresos

        if utilidad_anterior is None or utilidad_anterior == 0:
            variacion = None
        else:
            variacion = ((utilidad - utilidad_anterior) / abs(utilidad_anterior)) * 100

        resultado.append({
            "mes": mes,
            "nombre_mes": nombres[mes - 1],
            "ingresos": ingresos,
            "egresos": egresos,
            "utilidad": utilidad,
            "variacion_pct": variacion,
        })
        utilidad_anterior = utilidad

    return resultado


# ─────────────────────────── EXPORTAR EXCEL ──────────────────────

def exportar_excel_reporte(año=None, mes=None):
    """Genera ~/KyoGym/Reportes/Reportes_KyoGym.xlsx con hojas Ingresos, Egresos, Comparacion_Meses.
    Si una hoja ya existe la reemplaza."""
    try:
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl no está instalado. Ejecuta: pip install openpyxl")

    if año is None:
        año = date.today().year
    if mes is None:
        mes = date.today().month

    REPORTES_DIR.mkdir(parents=True, exist_ok=True)
    output_path = REPORTES_DIR / "Reportes_KyoGym.xlsx"

    # Cargar o crear workbook
    if output_path.exists():
        wb = load_workbook(output_path)
    else:
        wb = Workbook()
        # Eliminar hoja por defecto
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    def _write_sheet(wb, nombre, headers, rows):
        if nombre in wb.sheetnames:
            del wb[nombre]
        ws = wb.create_sheet(nombre)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(1, col, h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                ws.cell(ri, ci, val)
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    # ─── Hoja Ingresos ───
    if mes:
        fecha_desde_i = date(año, mes, 1)
        if mes == 12:
            fecha_hasta_i = date(año, 12, 31)
        else:
            fecha_hasta_i = date(año, mes + 1, 1) - timedelta(days=1)
    else:
        fecha_desde_i = date(año, 1, 1)
        fecha_hasta_i = date(año, 12, 31)

    ingresos = listar_ingresos(fecha_desde=fecha_desde_i, fecha_hasta=fecha_hasta_i)
    ing_rows = [(i["fecha"], i["cliente_nombre"], i["concepto"], i["metodo"], i["monto"])
                for i in ingresos]
    _write_sheet(wb, "Ingresos",
                 ["Fecha", "Cliente", "Concepto", "Método", "Monto"],
                 ing_rows)

    # ─── Hoja Egresos ───
    egresos = listar_egresos(fecha_desde=fecha_desde_i, fecha_hasta=fecha_hasta_i)
    eg_rows = [(e["fecha"], e["categoria"], e["descripcion"], e["proveedor"], e["metodo"], e["monto"])
               for e in egresos]
    _write_sheet(wb, "Egresos",
                 ["Fecha", "Categoría", "Descripción", "Proveedor", "Método", "Monto"],
                 eg_rows)

    # ─── Hoja Comparacion_Meses ───
    comparacion = obtener_comparacion_meses(año)
    comp_rows = []
    for c in comparacion:
        var = f"{c['variacion_pct']:+.1f}%" if c["variacion_pct"] is not None else "—"
        comp_rows.append((c["nombre_mes"], c["ingresos"], c["egresos"], c["utilidad"], var))
    # Totales
    tot_i = sum(c["ingresos"] for c in comparacion)
    tot_e = sum(c["egresos"] for c in comparacion)
    tot_u = tot_i - tot_e
    comp_rows.append(("Total año", tot_i, tot_e, tot_u, ""))
    _write_sheet(wb, "Comparacion_Meses",
                 ["Mes", "Ingresos", "Egresos", "Utilidad", "Variación %"],
                 comp_rows)

    wb.save(output_path)
    return str(output_path)


# ─────────────────────────── EXPORTAR PDF ────────────────────────

def exportar_pdf_reporte(año=None, mes=None):
    """Genera un PDF con ingresos, egresos y comparación de meses."""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
    except ImportError:
        raise ImportError("reportlab no está instalado. Ejecuta: pip install reportlab")

    if año is None:
        año = date.today().year
    if mes is None:
        mes = date.today().month

    REPORTES_DIR.mkdir(parents=True, exist_ok=True)
    nombre_mes_str = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                      "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"][mes - 1]
    output_path = REPORTES_DIR / f"Reporte_{año}_{nombre_mes_str}.pdf"

    doc = SimpleDocTemplate(str(output_path), pagesize=letter,
                            rightMargin=0.5 * inch, leftMargin=0.5 * inch,
                            topMargin=0.75 * inch, bottomMargin=0.75 * inch)
    styles = getSampleStyleSheet()
    story = []

    AZUL = colors.HexColor("#4472C4")
    BLANCO = colors.white
    GRIS = colors.HexColor("#f2f2f2")

    def _tabla_pdf(headers, rows, col_widths=None):
        data = [headers] + rows
        t = Table(data, colWidths=col_widths)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), AZUL),
            ("TEXTCOLOR", (0, 0), (-1, 0), BLANCO),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [BLANCO, GRIS]),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("PADDING", (0, 0), (-1, -1), 4),
        ]))
        return t

    # Título
    story.append(Paragraph(f"<b>KyoGym — Reporte {nombre_mes_str} {año}</b>",
                            styles["Title"]))
    story.append(Spacer(1, 0.2 * inch))

    # Resumen del mes
    resumen = obtener_resumen_mes(año, mes)
    story.append(Paragraph("<b>Resumen del mes</b>", styles["Heading2"]))
    res_data = [
        ["Ingresos", f"${resumen['ingresos_mes']:,.2f}"],
        ["Egresos", f"${resumen['egresos_mes']:,.2f}"],
        ["Utilidad", f"${resumen['utilidad']:,.2f}"],
        ["Membresías activas", str(resumen["membresias_activas"])],
        ["Cuentas por cobrar", str(resumen["cuentas_por_cobrar"])],
    ]
    story.append(_tabla_pdf(["Concepto", "Valor"], res_data, [3 * inch, 2 * inch]))
    story.append(Spacer(1, 0.25 * inch))

    # Ingresos del período
    fecha_desde = date(año, mes, 1)
    if mes == 12:
        fecha_hasta = date(año, 12, 31)
    else:
        fecha_hasta = date(año, mes + 1, 1) - timedelta(days=1)

    ingresos = listar_ingresos(fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)
    story.append(Paragraph("<b>Ingresos del período</b>", styles["Heading2"]))
    if ingresos:
        ing_rows = [(i["fecha"], i["cliente_nombre"][:20], i["concepto"][:20],
                     i["metodo"], f"${i['monto']:,.2f}") for i in ingresos]
        total_ing = sum(i["monto"] for i in ingresos)
        ing_rows.append(("", "", "", "TOTAL", f"${total_ing:,.2f}"))
        story.append(_tabla_pdf(["Fecha", "Cliente", "Concepto", "Método", "Monto"],
                                ing_rows,
                                [1.1 * inch, 1.8 * inch, 1.8 * inch, 1 * inch, 1 * inch]))
    else:
        story.append(Paragraph("Sin ingresos en el período.", styles["Normal"]))
    story.append(Spacer(1, 0.25 * inch))

    # Egresos del período
    egresos = listar_egresos(fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)
    story.append(Paragraph("<b>Egresos del período</b>", styles["Heading2"]))
    if egresos:
        eg_rows = [(e["fecha"], e["categoria"], e["descripcion"][:20],
                    e["metodo"], f"${e['monto']:,.2f}") for e in egresos]
        total_eg = sum(e["monto"] for e in egresos)
        eg_rows.append(("", "", "", "TOTAL", f"${total_eg:,.2f}"))
        story.append(_tabla_pdf(["Fecha", "Categoría", "Descripción", "Método", "Monto"],
                                eg_rows,
                                [1.1 * inch, 1.2 * inch, 2 * inch, 1 * inch, 1 * inch]))
    else:
        story.append(Paragraph("Sin egresos en el período.", styles["Normal"]))
    story.append(Spacer(1, 0.25 * inch))

    # Comparación mes a mes
    comparacion = obtener_comparacion_meses(año)
    story.append(Paragraph(f"<b>Comparación mes a mes — {año}</b>", styles["Heading2"]))
    comp_rows = []
    for c in comparacion:
        var = f"{c['variacion_pct']:+.1f}%" if c["variacion_pct"] is not None else "—"
        comp_rows.append((c["nombre_mes"], f"${c['ingresos']:,.2f}",
                          f"${c['egresos']:,.2f}", f"${c['utilidad']:,.2f}", var))
    tot_i = sum(c["ingresos"] for c in comparacion)
    tot_e = sum(c["egresos"] for c in comparacion)
    comp_rows.append(("Total año", f"${tot_i:,.2f}", f"${tot_e:,.2f}", f"${tot_i-tot_e:,.2f}", ""))
    story.append(_tabla_pdf(["Mes", "Ingresos", "Egresos", "Utilidad", "Variación %"],
                            comp_rows,
                            [1.3 * inch, 1.2 * inch, 1.2 * inch, 1.2 * inch, 1 * inch]))

    doc.build(story)
    return str(output_path)
