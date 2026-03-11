"""
Script alternativo para sincronizar con OneDrive usando autenticación de código de dispositivo.
Ideal para cuentas personales de Microsoft.
"""
import sqlite3
from pathlib import Path
from datetime import datetime
import json
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("openpyxl no está instalado. Ejecuta: pip install openpyxl")

try:
    import msal
    import requests
except ImportError:
    raise ImportError("msal no está instalado. Ejecuta: pip install msal requests")

from utils.constants import DB_PATH


# ==================== CONFIGURACIÓN ====================
# Archivo de configuración para credenciales
CONFIG_FILE = Path(__file__).parent / "onedrive_config_personal.json"

# Configuración por defecto para cuentas personales
# Usando el client_id de Microsoft Graph Explorer (aplicación pública)
DEFAULT_CONFIG = {
    "client_id": "1950a258-227b-4e31-a9cf-717495945fc2",  # Azure PowerShell public client (soporta Device Code Flow)
    "authority": "https://login.microsoftonline.com/consumers",
    "scope": [
        "https://graph.microsoft.com/Files.ReadWrite"
    ],
    "onedrive_folder": "/",
    "excel_filename": "gimnasio.xlsx"
}

# Archivo para almacenar el token cache
TOKEN_CACHE_FILE = Path(__file__).parent / "onedrive_token_cache.bin"


class OneDriveSyncPersonal:
    """Clase para sincronizar con OneDrive usando cuentas personales"""
    
    def __init__(self, config_path=None):
        """Inicializa el sincronizador"""
        self.config = self._load_config(config_path)
        self.access_token = None
        self.token_cache = self._load_token_cache()
        
    def _load_config(self, config_path=None):
        """Carga la configuración"""
        if config_path is None:
            config_path = CONFIG_FILE
            
        if config_path.exists():
            with open(config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        else:
            print(f"📝 Creando configuración por defecto...")
            self._create_default_config(config_path)
            return DEFAULT_CONFIG
    
    def _create_default_config(self, config_path):
        """Crea configuración por defecto"""
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(DEFAULT_CONFIG, f, indent=4, ensure_ascii=False)
    
    def _load_token_cache(self):
        """Carga el cache de tokens si existe"""
        cache = msal.SerializableTokenCache()
        if TOKEN_CACHE_FILE.exists():
            cache.deserialize(open(TOKEN_CACHE_FILE, "r").read())
        return cache
    
    def _save_token_cache(self):
        """Guarda el cache de tokens"""
        if self.token_cache.has_state_changed:
            with open(TOKEN_CACHE_FILE, "w") as f:
                f.write(self.token_cache.serialize())
    
    def authenticate(self):
        """Autentica con Device Code Flow (evita errores de redirect_uri)."""
        print("\n🔐 Autenticando con Microsoft (cuenta personal)...")

        app = msal.PublicClientApplication(
            self.config["client_id"],
            authority=self.config["authority"],
            token_cache=self.token_cache
        )

        # 1) Intentar token en caché
        accounts = app.get_accounts()
        if accounts:
            result = app.acquire_token_silent(self.config["scope"], account=accounts[0])
            if result and "access_token" in result:
                self.access_token = result["access_token"]
                self._save_token_cache()
                print("✅ Token recuperado desde caché")
                return True

        # 2) Device code flow (sin redirect_uri)
        return self._authenticate_device_flow(app)


    def _authenticate_device_flow(self, app):
        """Flujo de código de dispositivo"""
        flow = app.initiate_device_flow(scopes=self.config["scope"])
        if "user_code" not in flow:
            raise Exception(f"No se pudo iniciar device flow: {flow}")

        print("\n📱 Inicia sesión en:")
        print(flow["verification_uri"])
        print(f"Código: {flow['user_code']}\n")

        result = app.acquire_token_by_device_flow(flow)

        if "access_token" in result:
            self.access_token = result["access_token"]
            self._save_token_cache()
            print("✅ Autenticación exitosa")
            return True

        raise Exception(f"Error de autenticación: {result.get('error_description', result)}")
    
    def read_database(self):
        """Lee los datos de la base de datos"""
        print(f"\n📊 Leyendo datos de la base de datos...")
        
        if not DB_PATH.exists():
            raise FileNotFoundError(f"Base de datos no encontrada: {DB_PATH}")
        
        conn = sqlite3.connect(str(DB_PATH))
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        data = {}
        
        # Leer clientes
        cursor.execute("SELECT * FROM clientes ORDER BY id")
        data['clientes'] = [dict(row) for row in cursor.fetchall()]
        print(f"  ✓ Clientes: {len(data['clientes'])} registros")
        
        # Leer membresías
        cursor.execute("""
            SELECT m.*, c.nombre as cliente_nombre 
            FROM membresias m
            LEFT JOIN clientes c ON m.cliente_id = c.id
            ORDER BY m.id
        """)
        data['membresias'] = [dict(row) for row in cursor.fetchall()]
        print(f"  ✓ Membresías: {len(data['membresias'])} registros")
        
        # Leer pagos
        cursor.execute("""
            SELECT p.*, c.nombre as cliente_nombre 
            FROM pagos p
            LEFT JOIN clientes c ON p.cliente_id = c.id
            ORDER BY p.id
        """)
        data['pagos'] = [dict(row) for row in cursor.fetchall()]
        print(f"  ✓ Pagos: {len(data['pagos'])} registros")
        
        # Leer inventario
        cursor.execute("SELECT * FROM inventario ORDER BY id")
        data['inventario'] = [dict(row) for row in cursor.fetchall()]
        print(f"  ✓ Inventario: {len(data['inventario'])} registros")
        
        # Leer egresos
        try:
            cursor.execute("SELECT * FROM egresos ORDER BY id")
            data['egresos'] = [dict(row) for row in cursor.fetchall()]
            print(f"  ✓ Egresos: {len(data['egresos'])} registros")
        except Exception:
            data['egresos'] = []
            print("  ! Tabla egresos no encontrada (se omite)")
        
        conn.close()
        return data
    
    def create_excel(self, data, output_path):
        """Crea un archivo Excel con los datos"""
        print(f"\n📝 Creando archivo Excel...")
        
        wb = Workbook()
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Hoja 1: Resumen
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        
        ws_resumen['A1'] = "REPORTE GIMNASIO KYO-GYM"
        ws_resumen['A1'].font = Font(bold=True, size=16)
        ws_resumen['A3'] = "Fecha de generación:"
        ws_resumen['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws_resumen['A5'] = "Total Clientes:"
        ws_resumen['B5'] = len(data['clientes'])
        ws_resumen['A6'] = "Total Membresías:"
        ws_resumen['B6'] = len(data['membresias'])
        ws_resumen['A7'] = "Total Pagos:"
        ws_resumen['B7'] = len(data['pagos'])
        ws_resumen['A8'] = "Total Productos en Inventario:"
        ws_resumen['B8'] = len(data['inventario'])
        
        # Hoja 2: Clientes
        ws_clientes = wb.create_sheet("Clientes")
        
        if data['clientes']:
            headers = list(data['clientes'][0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws_clientes.cell(1, col, header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            for row_idx, cliente in enumerate(data['clientes'], 2):
                for col_idx, header in enumerate(headers, 1):
                    ws_clientes.cell(row_idx, col_idx, cliente[header])
            
            for col in range(1, len(headers) + 1):
                ws_clientes.column_dimensions[get_column_letter(col)].width = 15
        
        # Hoja 3: Membresías
        ws_membresias = wb.create_sheet("Membresías")
        
        if data['membresias']:
            headers = list(data['membresias'][0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws_membresias.cell(1, col, header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            for row_idx, membresia in enumerate(data['membresias'], 2):
                for col_idx, header in enumerate(headers, 1):
                    ws_membresias.cell(row_idx, col_idx, membresia[header])
            
            for col in range(1, len(headers) + 1):
                ws_membresias.column_dimensions[get_column_letter(col)].width = 15
        
        # Hoja 4: Pagos
        ws_pagos = wb.create_sheet("Pagos")
        
        if data['pagos']:
            headers = list(data['pagos'][0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws_pagos.cell(1, col, header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            for row_idx, pago in enumerate(data['pagos'], 2):
                for col_idx, header in enumerate(headers, 1):
                    ws_pagos.cell(row_idx, col_idx, pago[header])
            
            for col in range(1, len(headers) + 1):
                ws_pagos.column_dimensions[get_column_letter(col)].width = 15
        
        # Hoja 5: Inventario
        ws_inventario = wb.create_sheet("Inventario")
        
        if data['inventario']:
            headers = list(data['inventario'][0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws_inventario.cell(1, col, header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            for row_idx, producto in enumerate(data['inventario'], 2):
                for col_idx, header in enumerate(headers, 1):
                    ws_inventario.cell(row_idx, col_idx, producto[header])
            
            for col in range(1, len(headers) + 1):
                ws_inventario.column_dimensions[get_column_letter(col)].width = 15
        
        # Hoja 6: Egresos
        ws_egresos = wb.create_sheet("Egresos")
        
        if data.get('egresos'):
            headers = list(data['egresos'][0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws_egresos.cell(1, col, header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            for row_idx, egreso in enumerate(data['egresos'], 2):
                for col_idx, header in enumerate(headers, 1):
                    ws_egresos.cell(row_idx, col_idx, egreso[header])
            
            for col in range(1, len(headers) + 1):
                ws_egresos.column_dimensions[get_column_letter(col)].width = 15
        
        wb.save(output_path)
        print(f"✅ Archivo Excel creado")
        
        return output_path
    
    def upload_to_onedrive(self, file_path):
        """Sube el archivo a OneDrive"""
        print(f"\n☁️  Subiendo archivo a OneDrive...")
        
        if not self.access_token:
            raise Exception("No hay token de acceso")
        
        filename = self.config.get("excel_filename", "gimnasio.xlsx")
        onedrive_folder = self.config.get("onedrive_folder", "/")
        
        # Construir URL
        if onedrive_folder == "/":
            upload_path = f"/me/drive/root:/{filename}:/content"
        else:
            upload_path = f"/me/drive/root:{onedrive_folder}/{filename}:/content"
        
        url = f"https://graph.microsoft.com/v1.0{upload_path}"
        
        headers = {
            "Authorization": f"Bearer {self.access_token}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
        
        with open(file_path, 'rb') as f:
            file_data = f.read()
        
        try:
            response = requests.put(url, headers=headers, data=file_data)
            response.raise_for_status()
            
            result = response.json()
            print(f"✅ Archivo subido exitosamente")
            print(f"   📁 Nombre: {result.get('name')}")
            print(f"   📏 Tamaño: {result.get('size')} bytes")
            
            if 'webUrl' in result:
                print(f"   🌐 URL: {result.get('webUrl')}")
            
            return result
            
        except requests.exceptions.HTTPError as e:
            error_detail = e.response.json() if e.response.text else str(e)
            print(f"❌ Error al subir: {error_detail}")
            raise
    
    def _get_googledrive_local_path(self):
        """Detecta la carpeta local de Google Drive automáticamente"""
        candidates = [
            # Google Drive for Desktop (nuevo cliente)
            Path.home() / "My Drive",
            Path("G:\\") / "My Drive",
            Path("G:\\"),
            # Google Drive (cliente antiguo)
            Path.home() / "Google Drive",
            Path("G:\\Google Drive"),
        ]

        # Buscar en todas las letras de unidad (G:, H:, etc.)
        import string
        for letter in string.ascii_uppercase:
            drive = Path(f"{letter}:\\")
            try:
                if not drive.exists():
                    continue
            except OSError:
                continue
            try:
                # Primero buscar subcarpeta (en inglés o español)
                for sub in ["Mi unidad", "My Drive", "Google Drive"]:
                    p = drive / sub
                    try:
                        if p.exists():
                            return p
                    except OSError:
                        continue
                # Si la raíz tiene el marcador de Google Drive
                marker = drive / ".shortcut-targets-by-id"
                try:
                    if marker.exists():
                        # Crear "My Drive" si no existe
                        my_drive = drive / "My Drive"
                        try:
                            my_drive.mkdir(exist_ok=True)
                        except OSError:
                            pass
                        if my_drive.exists():
                            return my_drive
                except OSError:
                    pass
            except OSError:
                continue

        for p in candidates:
            try:
                if p.exists():
                    return p
            except OSError:
                continue
        return None

    def sync(self):
        """Ejecuta la sincronización completa guardando en carpeta local de Google Drive"""
        print("=" * 60)
        print("🔄 SINCRONIZACIÓN GIMNASIO.DB → GOOGLE DRIVE (Carpeta Local)")
        print("=" * 60)

        try:
            # Detectar carpeta Google Drive local
            gdrive_path = self._get_googledrive_local_path()
            if not gdrive_path:
                raise FileNotFoundError(
                    "No se encontró la carpeta local de Google Drive.\n"
                    "   1. Descarga e instala 'Google Drive for Desktop':\n"
                    "      https://www.google.com/drive/download/\n"
                    "   2. Inicia sesión con kyogymdata@gmail.com\n"
                    "   3. Vuelve a ejecutar este script."
                )

            filename = self.config.get("excel_filename", "gimnasio.xlsx")
            output_file = gdrive_path / filename

            print(f"\n📁 Carpeta Google Drive detectada: {gdrive_path}")

            # Leer datos
            data = self.read_database()

            # Crear Excel directamente en la carpeta de Google Drive
            self.create_excel(data, output_file)

            print(f"\n☁️  Archivo guardado en: {output_file}")
            print("   Google Drive sincronizará automáticamente con la cuenta kyogymdata@gmail.com")

            print("\n" + "=" * 60)
            print("✅ SINCRONIZACIÓN COMPLETADA")
            print(f"⏰ {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            print("=" * 60)

            return True

        except Exception as e:
            print(f"\n❌ Error: {str(e)}")
            return False


def main():
    """Función principal"""
    print("\n🏋️ KyoGym - Sincronización con OneDrive Personal")
    print("=" * 60 + "\n")
    
    try:
        syncer = OneDriveSyncPersonal()
        syncer.sync()
        
    except KeyboardInterrupt:
        print("\n\n⚠️  Cancelado por el usuario")
    except Exception as e:
        print(f"\n❌ Error fatal: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
