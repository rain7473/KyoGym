"""Vista de configuración del sistema"""
from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel,
                               QFrame, QLineEdit, QPushButton, QGroupBox,
                               QFormLayout, QComboBox, QSpinBox, QDialog,
                               QMessageBox, QScrollArea, QFileDialog,
                               QTableWidget, QTableWidgetItem, QHeaderView,
                               QTabWidget, QDateEdit, QSizePolicy, QGridLayout)
from PySide6.QtCore import Qt, Signal, QDate
from PySide6.QtGui import QFont, QPixmap, QColor
from utils.iconos_ui import crear_boton_icono
from utils.table_styles import aplicar_estilo_tabla_moderna
from utils.table_utils import limpiar_tabla
from utils.validators import crear_validador_nombre, TelefonoFormateadoLineEdit, crear_validador_email
from usuario_activo import obtener_usuario_activo
from db import create_user, get_all_users, delete_user
from services import auditoria_service
import json
import os
from collections import defaultdict


DEFAULT_DIAS_ALERTA_VENCIMIENTO = 7

_MSG_SS = """
    QMessageBox { background-color: #ffffff; }
    QLabel { color: #1a1a1a; font-size: 13px; min-width: 300px; border: none; }
    QPushButton {
        background-color: #3498db; color: white;
        padding: 8px 20px; border: none; border-radius: 4px;
        font-weight: bold; font-size: 13px; min-width: 80px;
    }
    QPushButton:hover { background-color: #2980b9; }
"""


class VerUsuariosDialog(QDialog):
    """Diálogo que lista todos los usuarios con opción de eliminar por fila."""

    def __init__(self, usuario_activo="", parent=None):
        super().__init__(parent)
        self.usuario_activo = usuario_activo
        self.setWindowTitle("Usuarios del sistema")
        self.setMinimumSize(580, 380)
        self._init_ui()
        self._cargar()

    def _init_ui(self):
        self.setStyleSheet("""
            QDialog { background-color: #f5f5f5; }
            QLabel { color: #1a1a1a; font-size: 13px; border: none; background: transparent; }
            QPushButton#btn_cerrar {
                background-color: #7f8c8d; color: white;
                padding: 8px 20px; border: none;
                border-radius: 4px; font-weight: bold; font-size: 13px;
            }
            QPushButton#btn_cerrar:hover { background-color: #636e72; }
        """)

        layout = QVBoxLayout(self)
        layout.setSpacing(14)

        titulo = QLabel("Usuarios registrados")
        titulo.setFont(QFont("Arial", 15, QFont.Bold))
        titulo.setStyleSheet("color: #1a1a1a; margin-bottom: 4px; border: none; background: transparent;")
        layout.addWidget(titulo)

        self.tabla = QTableWidget(0, 4)
        self.tabla.setHorizontalHeaderLabels(["Nombre completo", "Usuario", "Rol", "Acción"])
        self.tabla.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.tabla.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.tabla.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.tabla.horizontalHeader().setSectionResizeMode(3, QHeaderView.Fixed)
        self.tabla.setColumnWidth(3, 110)
        self.tabla.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tabla.setSelectionMode(QTableWidget.NoSelection)
        self.tabla.verticalHeader().setVisible(False)
        aplicar_estilo_tabla_moderna(self.tabla, compacta=True)
        layout.addWidget(self.tabla)

        btn_layout = QHBoxLayout()
        btn_cerrar = QPushButton("Cerrar")
        btn_cerrar.setObjectName("btn_cerrar")
        btn_cerrar.clicked.connect(self.accept)
        btn_layout.addWidget(btn_cerrar)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)

    def _cargar(self):
        usuarios = get_all_users()
        limpiar_tabla(self.tabla)
        self.tabla.setRowCount(len(usuarios))
        for row, u in enumerate(usuarios):
            username = u['username']
            self.tabla.setItem(row, 0, QTableWidgetItem(u.get('full_name') or username))
            self.tabla.setItem(row, 1, QTableWidgetItem(username))
            self.tabla.setItem(row, 2, QTableWidgetItem(u['role']))

            btn_del = crear_boton_icono("delete.svg", "#e74c3c", "#c0392b", "Eliminar usuario")
            if username == self.usuario_activo:
                btn_del.setEnabled(False)
                btn_del.setToolTip("No puedes eliminar tu propio usuario")
            else:
                btn_del.clicked.connect(lambda checked, usr=username: self._confirmar_eliminar(usr))
            self.tabla.setCellWidget(row, 3, btn_del)

    def _confirmar_eliminar(self, username):
        msg = QMessageBox(self)
        msg.setWindowTitle("Confirmar")
        msg.setText(f"¿Eliminar el usuario \'{username}\'?\nEsta acción no se puede deshacer.")
        msg.setStyleSheet("""
            QMessageBox { background-color: #f5f5f5; }
            QLabel { color: #2c2c2c; font-size: 14px; min-width: 280px; border: none; }
            QPushButton {
                background-color: #2c3e50; color: white;
                padding: 8px 20px; border: none;
                border-radius: 4px; font-weight: bold;
                font-size: 13px; min-width: 70px;
            }
            QPushButton:hover { background-color: #3d5166; }
        """)
        btn_si = msg.addButton("Sí", QMessageBox.YesRole)
        btn_no = msg.addButton("No", QMessageBox.NoRole)
        msg.setDefaultButton(btn_no)
        msg.exec()
        if msg.clickedButton() != btn_si:
            return
        ok = delete_user(username)
        if ok:
            self._cargar()
        else:
            QMessageBox.critical(self, "Error", "No se pudo eliminar el usuario.")


class CrearUsuarioDialog(QDialog):
    """Diálogo para crear un nuevo usuario del sistema."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Crear nuevo usuario")
        self.setMinimumWidth(420)
        self._init_ui()

    def _init_ui(self):
        self.setStyleSheet("""
            QDialog { background-color: #f5f5f5; }
            QLabel { color: #1a1a1a; font-size: 13px; }
            QLineEdit, QComboBox {
                padding: 7px; font-size: 13px; color: #1a1a1a;
                border: 2px solid #d0d0d0; border-radius: 5px;
                background-color: #f5f5f5;
            }
            QPushButton {
                background-color: #27ae60; color: white;
                padding: 10px 24px; border: none;
                border-radius: 5px; font-weight: bold; font-size: 13px;
            }
            QPushButton:hover { background-color: #229954; }
        """)

        layout = QVBoxLayout(self)
        layout.setSpacing(14)

        titulo = QLabel("Nuevo usuario")
        titulo.setFont(QFont("Arial", 15, QFont.Bold))
        titulo.setStyleSheet("color: #1a1a1a; margin-bottom: 6px;")
        layout.addWidget(titulo)

        form = QFormLayout()
        form.setSpacing(10)

        self.txt_username = QLineEdit()
        self.txt_username.setPlaceholderText("Nombre de usuario")
        form.addRow("Usuario:", self.txt_username)

        self.txt_nombre = QLineEdit()
        self.txt_nombre.setPlaceholderText("Nombre completo (opcional)")
        form.addRow("Nombre completo:", self.txt_nombre)

        self.txt_password = QLineEdit()
        self.txt_password.setEchoMode(QLineEdit.Password)
        self.txt_password.setPlaceholderText("Contraseña")
        form.addRow("Contraseña:", self.txt_password)

        self.txt_password2 = QLineEdit()
        self.txt_password2.setEchoMode(QLineEdit.Password)
        self.txt_password2.setPlaceholderText("Repetir contraseña")
        form.addRow("Confirmar:", self.txt_password2)

        self.cmb_rol = QComboBox()
        self.cmb_rol.addItems(["user", "admin"])
        self.cmb_rol.setEditable(False)
        form.addRow("Rol:", self.cmb_rol)

        layout.addLayout(form)

        btn_crear = QPushButton("➕ Crear usuario")
        btn_crear.clicked.connect(self._crear)
        layout.addWidget(btn_crear)

    def _crear(self):
        username = self.txt_username.text().strip()
        nombre = self.txt_nombre.text().strip() or username
        password = self.txt_password.text()
        password2 = self.txt_password2.text()
        role = self.cmb_rol.currentText()

        if not username or not password:
            QMessageBox.warning(self, "Error", "Usuario y contraseña son obligatorios.")
            return
        if password != password2:
            QMessageBox.warning(self, "Error", "Las contraseñas no coinciden.")
            return

        ok = create_user(username, password, full_name=nombre, role=role)
        if ok:
            QMessageBox.information(self, "Éxito",
                f"Usuario \'{username}\' creado correctamente con rol \'{role}\'.")
            self.accept()
        else:
            QMessageBox.critical(self, "Error",
                "No se pudo crear el usuario. Es posible que ya exista.")


# ───────────────────────────────────────────────────────────────────────────
# Colores de avatar compartidos
# ───────────────────────────────────────────────────────────────────────────
_AVATAR_PALETTE = [
    "#2563eb", "#16a34a", "#dc2626", "#7c3aed", "#db2777",
    "#0891b2", "#d97706", "#059669", "#9333ea", "#e11d48",
]

_COLORES_MOD_BADGE = {
    "Clientes":   ("#eaf4ff", "#1558a8"),
    "Pagos":      ("#eafff2", "#1a6b3a"),
    "Membresías": ("#fffbe6", "#8a5c00"),
    "Inventario": ("#f5eeff", "#5c2d91"),
}

_RANK_HEADER_COLOR = {1: "#f59e0b", 2: "#94a3b8", 3: "#b45309"}


class PerfilEmpleadoDialog(QDialog):
    """Muestra el perfil detallado de un usuario con sus estadísticas de actividad."""

    def __init__(self, udata: dict, modulos_data: dict,
                 ranking: int, total_ranking: int, parent=None):
        super().__init__(parent)
        self._udata         = udata           # {username, full_name, role}
        self._modulos       = modulos_data    # {modulo: count}
        self._ranking       = ranking         # 1-based
        self._total_ranking = total_ranking
        nombre = udata.get('full_name') or udata.get('username', 'Usuario')
        self.setWindowTitle(f"Perfil de actividad — {nombre}")
        self.setMinimumWidth(620)
        self.setMinimumHeight(580)
        self.setModal(True)
        self._init_ui()

    def _init_ui(self):
        self.setStyleSheet("QDialog { background-color: #f5f7fa; }")
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Cabecera coloreada según rango ──────────────────────────────
        rank_color = _RANK_HEADER_COLOR.get(self._ranking, "#2c3e50")
        header = QFrame()
        header.setFixedHeight(104)
        header.setStyleSheet(f"background-color: {rank_color}; border: none;")
        h_lay = QHBoxLayout(header)
        h_lay.setContentsMargins(24, 14, 24, 14)
        h_lay.setSpacing(16)

        username  = self._udata.get('username', '')
        full_name = self._udata.get('full_name') or username
        initials  = (full_name[:1] if full_name else "?").upper()
        av_color  = _AVATAR_PALETTE[sum(ord(c) for c in username) % len(_AVATAR_PALETTE)]

        avatar = QLabel(initials)
        avatar.setFixedSize(64, 64)
        avatar.setAlignment(Qt.AlignCenter)
        avatar.setStyleSheet(f"""
            background-color: {av_color}; color: white;
            font-size: 24px; font-weight: bold; border-radius: 32px;
            border: 3px solid rgba(255,255,255,0.55);
        """)
        h_lay.addWidget(avatar)

        info_l = QVBoxLayout()
        info_l.setSpacing(3)
        lbl_nm = QLabel(full_name)
        lbl_nm.setFont(QFont("Arial", 15, QFont.Bold))
        lbl_nm.setStyleSheet("color: white; background: transparent; border: none;")
        lbl_mt = QLabel(f"@{username}  ·  {self._udata.get('role', '—')}")
        lbl_mt.setStyleSheet(
            "color: rgba(255,255,255,0.82); font-size: 12px; background: transparent; border: none;"
        )
        info_l.addWidget(lbl_nm)
        info_l.addWidget(lbl_mt)
        h_lay.addLayout(info_l, 1)

        # Medalla de rango
        medals = {1: "🥇\n#1", 2: "🥈\n#2", 3: "🥉\n#3"}
        rank_txt = medals.get(self._ranking, f"#{self._ranking}")
        lbl_rank = QLabel(rank_txt)
        lbl_rank.setFont(QFont("Arial", 18 if self._ranking <= 3 else 16, QFont.Bold))
        lbl_rank.setAlignment(Qt.AlignCenter)
        lbl_rank.setStyleSheet("color: white; background: transparent; border: none;")
        h_lay.addWidget(lbl_rank)

        root.addWidget(header)

        # ── Cuerpo ──────────────────────────────────────────────────────
        body = QWidget()
        body.setStyleSheet("background-color: #f5f7fa;")
        b_lay = QVBoxLayout(body)
        b_lay.setContentsMargins(24, 20, 24, 20)
        b_lay.setSpacing(16)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setStyleSheet("""
            QScrollArea { border: none; background-color: #f5f7fa; }
            QScrollBar:vertical {
                background: #f0f0f0; width: 8px; border-radius: 4px;
            }
            QScrollBar::handle:vertical {
                background: #c0c0c0; border-radius: 4px; min-height: 26px;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }
        """)
        scroll.setWidget(body)
        root.addWidget(scroll)

        # KPI total acciones
        total_acc = sum(self._modulos.values())
        kpi = QFrame()
        kpi.setStyleSheet(
            "QFrame { background-color: white; border: 1px solid #e2e8f0; border-radius: 10px; }"
        )
        kpi_lay = QHBoxLayout(kpi)
        kpi_lay.setContentsMargins(20, 14, 20, 14)
        kpi_lay.setSpacing(14)

        ico = QLabel("⚡")
        ico.setFont(QFont("Arial", 22))
        ico.setStyleSheet("background: transparent; border: none;")
        kpi_lay.addWidget(ico)

        nums = QVBoxLayout()
        nums.setSpacing(2)
        lbl_n = QLabel(f"{total_acc:,}")
        lbl_n.setFont(QFont("Arial", 28, QFont.Bold))
        lbl_n.setStyleSheet("color: #1a2332; background: transparent; border: none;")
        lbl_d = QLabel("acciones totales registradas")
        lbl_d.setStyleSheet(
            "color: #64748b; font-size: 13px; background: transparent; border: none;"
        )
        plural = "s" if self._total_ranking != 1 else ""
        lbl_p = QLabel(f"Puesto #{self._ranking} de {self._total_ranking} usuario{plural}")
        lbl_p.setStyleSheet(
            "color: #94a3b8; font-size: 12px; background: transparent; border: none;"
        )
        nums.addWidget(lbl_n)
        nums.addWidget(lbl_d)
        nums.addWidget(lbl_p)
        kpi_lay.addLayout(nums, 1)
        b_lay.addWidget(kpi)

        # Desglose por módulo
        if self._modulos:
            lbl_sec = QLabel("Actividad por módulo")
            lbl_sec.setFont(QFont("Arial", 13, QFont.Bold))
            lbl_sec.setStyleSheet(
                "color: #1a2332; background: transparent; border: none;"
            )
            b_lay.addWidget(lbl_sec)

            mod_card = QFrame()
            mod_card.setStyleSheet(
                "QFrame { background-color: white; border: 1px solid #e2e8f0; border-radius: 10px; }"
            )
            mc_lay = QVBoxLayout(mod_card)
            mc_lay.setContentsMargins(20, 14, 20, 14)
            mc_lay.setSpacing(10)

            max_c = max(self._modulos.values())
            for mod, cnt in sorted(self._modulos.items(), key=lambda x: -x[1]):
                bg, fg = _COLORES_MOD_BADGE.get(mod, ("#f5f7fa", "#1a2332"))
                row = QHBoxLayout()
                row.setSpacing(12)

                lbl_mod = QLabel(mod)
                lbl_mod.setFixedWidth(95)
                lbl_mod.setAlignment(Qt.AlignCenter)
                lbl_mod.setStyleSheet(
                    f"color: {fg}; background-color: {bg}; font-size: 12px; font-weight: bold;"
                    " padding: 3px 8px; border-radius: 4px; border: none;"
                )
                row.addWidget(lbl_mod)

                bar_outer = QFrame()
                bar_outer.setFixedHeight(14)
                bar_outer.setFixedWidth(200)
                bar_outer.setStyleSheet(
                    "background-color: #f1f5f9; border-radius: 7px; border: none;"
                )
                bar_inner = QFrame(bar_outer)
                bar_inner.setFixedHeight(14)
                bar_inner.setFixedWidth(max(12, int((cnt / max_c) * 200)))
                bar_inner.setStyleSheet(
                    f"background-color: {fg}; border-radius: 7px; border: none;"
                )
                row.addWidget(bar_outer)

                lbl_cnt = QLabel(f"{cnt} acción{'es' if cnt != 1 else ''}")
                lbl_cnt.setStyleSheet(
                    "color: #475569; font-size: 13px; font-weight: bold;"
                    " background: transparent; border: none;"
                )
                row.addWidget(lbl_cnt)
                row.addStretch()
                mc_lay.addLayout(row)

            b_lay.addWidget(mod_card)

        # ── Historial reciente ──────────────────────────────────────────
        lbl_hist = QLabel("Historial reciente")
        lbl_hist.setFont(QFont("Arial", 13, QFont.Bold))
        lbl_hist.setStyleSheet(
            "color: #1a2332; background: transparent; border: none;"
        )
        b_lay.addWidget(lbl_hist)

        hist_card = QFrame()
        hist_card.setStyleSheet(
            "QFrame { background-color: white; border: 1px solid #e2e8f0; border-radius: 10px; }"
        )
        hc_lay = QVBoxLayout(hist_card)
        hc_lay.setContentsMargins(0, 0, 0, 0)
        hc_lay.setSpacing(0)

        hist_table = QTableWidget(0, 4)
        hist_table.setHorizontalHeaderLabels(["Fecha", "Módulo", "Acción", "Descripción"])
        hist_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        hist_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        hist_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        hist_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        hist_table.setEditTriggers(QTableWidget.NoEditTriggers)
        hist_table.setSelectionMode(QTableWidget.NoSelection)
        hist_table.verticalHeader().setVisible(False)
        hist_table.setMinimumHeight(180)
        hist_table.setMaximumHeight(280)
        aplicar_estilo_tabla_moderna(hist_table, compacta=True)

        _username = self._udata.get('username', '')
        _todos = auditoria_service.obtener_historial(limite=500)
        _filas = [r for r in _todos if r.get('usuario', '') == _username][:50]

        _BADGE_COLORES = {
            "CREAR":     ("#1abc9c", "#ffffff"),
            "MODIFICAR": ("#f39c12", "#ffffff"),
            "EDITAR":    ("#f39c12", "#ffffff"),
            "ELIMINAR":  ("#e74c3c", "#ffffff"),
            "PAGO":      ("#e67e22", "#ffffff"),
            "IMPORTAR":  ("#9b59b6", "#ffffff"),
            "RENOVAR":   ("#2980b9", "#ffffff"),
        }

        hist_table.setRowCount(len(_filas))
        for row_i, reg in enumerate(_filas):
            fecha_raw = reg.get('fecha_hora', '')
            fecha_str = fecha_raw[:16].replace('T', ' ') if fecha_raw else '—'

            modulo   = reg.get('modulo', '—')
            accion   = reg.get('accion', '—')
            desc     = reg.get('descripcion', '')

            item_f = QTableWidgetItem(fecha_str)
            item_f.setTextAlignment(Qt.AlignCenter)

            item_mod = QTableWidgetItem(modulo)
            item_mod.setTextAlignment(Qt.AlignCenter)
            bg_mod, fg_mod = _COLORES_MOD_BADGE.get(modulo, ("#f5f7fa", "#1a2332"))
            item_mod.setBackground(QColor(bg_mod))
            item_mod.setForeground(QColor(fg_mod))

            item_acc = QTableWidgetItem(accion)
            item_acc.setTextAlignment(Qt.AlignCenter)
            bg_acc, fg_acc = _BADGE_COLORES.get(accion.upper(), ("#7f8c8d", "#ffffff"))
            item_acc.setBackground(QColor(bg_acc))
            item_acc.setForeground(QColor(fg_acc))

            item_desc = QTableWidgetItem(desc)

            hist_table.setItem(row_i, 0, item_f)
            hist_table.setItem(row_i, 1, item_mod)
            hist_table.setItem(row_i, 2, item_acc)
            hist_table.setItem(row_i, 3, item_desc)

        if not _filas:
            hist_table.setRowCount(1)
            item_vacio = QTableWidgetItem("Sin registros de actividad")
            item_vacio.setTextAlignment(Qt.AlignCenter)
            item_vacio.setForeground(QColor("#94a3b8"))
            hist_table.setSpan(0, 0, 1, 4)
            hist_table.setItem(0, 0, item_vacio)

        hc_lay.addWidget(hist_table)
        b_lay.addWidget(hist_card)

        # Botón cerrar
        btn_cerrar = QPushButton("Cerrar")
        btn_cerrar.setCursor(Qt.PointingHandCursor)
        btn_cerrar.setStyleSheet("""
            QPushButton {
                background-color: #2c3e50; color: white;
                padding: 10px 28px; border: none;
                border-radius: 8px; font-size: 13px; font-weight: bold;
            }
            QPushButton:hover { background-color: #3d5166; }
        """)
        btn_cerrar.clicked.connect(self.accept)
        br = QHBoxLayout()
        br.addStretch()
        br.addWidget(btn_cerrar)
        b_lay.addLayout(br)


class ConfiguracionView(QWidget):
    """Vista de configuración del sistema"""
    logout_solicitado = Signal()

    def __init__(self):
        super().__init__()
        self.config_file = os.path.join(os.path.dirname(os.path.dirname(__file__)), "config.json")
        self.logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "assets", "logo.png")
        self._usuario_activo = ""
        self._usuario_role   = ""
        # Paginación del historial
        self._hist_todos       = []
        self._hist_pagina      = 0
        self._hist_por_pagina  = 10
        self.init_ui()
        self.cargar_configuracion()
    
    def init_ui(self):
        """Inicializa la interfaz de usuario con pestañas."""
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        self.setStyleSheet("QLabel { color: #1a1a1a; }")

        # ── Encabezado ──────────────────────────────────────────────────
        header = QWidget()
        header.setStyleSheet("background-color: #ffffff; border-bottom: 1px solid #e0e0e0;")
        header.setFixedHeight(64)
        hlay = QHBoxLayout(header)
        hlay.setContentsMargins(28, 0, 28, 0)

        icono = QLabel("⚙️")
        icono.setFont(QFont("Arial", 22))
        icono.setStyleSheet("background: transparent; border: none;")
        hlay.addWidget(icono)

        titulo_header = QLabel("Configuración del Sistema")
        titulo_header.setFont(QFont("Arial", 18, QFont.Bold))
        titulo_header.setStyleSheet("color: #1a1a1a; background: transparent; border: none;")
        hlay.addWidget(titulo_header)
        hlay.addStretch()
        main_layout.addWidget(header)

        # ── Pestañas ─────────────────────────────────────────────────────
        self._tabs = QTabWidget()
        self._tabs.setDocumentMode(True)
        self._tabs.setStyleSheet("""
            QTabWidget::pane {
                border: none;
                background-color: #f5f7fa;
            }
            QTabBar::tab {
                background: transparent;
                color: #888888;
                padding: 12px 28px;
                font-size: 14px;
                font-weight: bold;
                border: none;
                border-bottom: 3px solid transparent;
            }
            QTabBar::tab:selected {
                color: #2c3e50;
                border-bottom: 3px solid #2c3e50;
                background: transparent;
            }
            QTabBar::tab:hover:!selected {
                color: #555555;
                border-bottom: 3px solid #cccccc;
            }
        """)

        self._tabs.addTab(self._crear_tab_general(),   "  General")
        self._tabs.addTab(self._crear_tab_historial(), "  Historial de acciones")
        self._tabs.addTab(self._crear_tab_resumen(),   "  Resumen")
        main_layout.addWidget(self._tabs)

    def _crear_tab_general(self):
        """Crea la pestaña de configuración general con diseño moderno en cards."""
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setStyleSheet("""
            QScrollArea { border: none; background-color: #f5f7fa; }
            QScrollBar:vertical {
                background: #f0f0f0; width: 8px; border-radius: 4px;
            }
            QScrollBar::handle:vertical {
                background: #c0c0c0; border-radius: 4px; min-height: 26px;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }
        """)

        container = QWidget()
        container.setStyleSheet("background-color: #f5f7fa;")
        scroll.setWidget(container)

        layout = QVBoxLayout(container)
        layout.setContentsMargins(32, 28, 32, 28)
        layout.setSpacing(20)

        # ── Fila superior: info del gimnasio (izq) + logo (der) ─────────
        top_row = QHBoxLayout()
        top_row.setSpacing(20)
        top_row.addWidget(self.crear_grupo_gimnasio(), stretch=3)
        top_row.addWidget(self.crear_grupo_logo(),     stretch=2)
        layout.addLayout(top_row)

        # ── Gestión de usuarios (visible solo para admin) ────────────────
        self._grupo_usuarios = self._crear_grupo_usuarios()
        self._grupo_usuarios.setVisible(False)
        layout.addWidget(self._grupo_usuarios)

        # ── Barra de acciones ────────────────────────────────────────────
        layout.addWidget(self._crear_barra_acciones())
        layout.addStretch()

        return scroll

    def _crear_barra_acciones(self):
        """Crea la barra inferior con los botones de acción principales."""
        bar = QFrame()
        bar.setObjectName("actionsBar")
        bar.setStyleSheet("""
            QFrame#actionsBar {
                background-color: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 12px;
            }
        """)
        hlay = QHBoxLayout(bar)
        hlay.setContentsMargins(24, 16, 24, 16)
        hlay.setSpacing(12)

        btn_guardar = QPushButton("💾  Guardar configuración")
        btn_guardar.setCursor(Qt.PointingHandCursor)
        btn_guardar.setStyleSheet("""
            QPushButton {
                background-color: #16a34a; color: white;
                padding: 12px 28px; border: none;
                border-radius: 8px; font-size: 14px; font-weight: bold;
            }
            QPushButton:hover { background-color: #15803d; }
        """)
        btn_guardar.clicked.connect(self.guardar_configuracion)

        btn_reset = QPushButton("🔄  Restaurar valores")
        btn_reset.setCursor(Qt.PointingHandCursor)
        btn_reset.setStyleSheet("""
            QPushButton {
                background-color: #f1f5f9; color: #64748b;
                padding: 12px 20px; border: 1.5px solid #e2e8f0;
                border-radius: 8px; font-size: 14px; font-weight: bold;
            }
            QPushButton:hover {
                background-color: #fee2e2; color: #dc2626; border-color: #fca5a5;
            }
        """)
        btn_reset.clicked.connect(self.restaurar_predeterminados)

        hlay.addWidget(btn_guardar)
        hlay.addWidget(btn_reset)
        hlay.addStretch()

        btn_cerrar = QPushButton("🔒  Cerrar sesión")
        btn_cerrar.setCursor(Qt.PointingHandCursor)
        btn_cerrar.setStyleSheet("""
            QPushButton {
                background-color: #f1f5f9; color: #64748b;
                padding: 12px 20px; border: 1.5px solid #e2e8f0;
                border-radius: 8px; font-size: 14px; font-weight: bold;
            }
            QPushButton:hover { background-color: #e2e8f0; color: #1a2332; }
        """)
        btn_cerrar.clicked.connect(self.cerrar_sesion)
        hlay.addWidget(btn_cerrar)

        return bar

    def _crear_grupo_usuarios(self):
        """Crea la card de gestión de usuarios (visible solo para admin)."""
        _CARD_SS = """
            QFrame#card {
                background-color: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 12px;
            }
        """
        card = QFrame()
        card.setObjectName("card")
        card.setStyleSheet(_CARD_SS)

        outer = QVBoxLayout(card)
        outer.setContentsMargins(24, 20, 24, 24)
        outer.setSpacing(14)

        # ── Cabecera ─────────────────────────────────────────────────────
        hdr = QHBoxLayout()
        ico = QLabel("👥")
        ico.setFont(QFont("Arial", 16))
        ico.setStyleSheet("background: transparent; border: none;")
        lbl_titulo = QLabel("Gestión de Usuarios")
        lbl_titulo.setFont(QFont("Arial", 14, QFont.Bold))
        lbl_titulo.setStyleSheet("color: #1a2332; background: transparent; border: none;")
        hdr.addWidget(ico)
        hdr.addSpacing(6)
        hdr.addWidget(lbl_titulo)
        hdr.addStretch()
        outer.addLayout(hdr)

        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet("background-color: #f0f4f8; border: none; max-height: 1px;")
        outer.addWidget(sep)

        lbl_desc = QLabel("Administra los usuarios que pueden acceder al sistema.")
        lbl_desc.setStyleSheet("color: #64748b; font-size: 13px; background: transparent; border: none;")
        outer.addWidget(lbl_desc)

        # ── Botones de acción ──────────────────────────────────────────────
        btn_crear = QPushButton("➕  Crear nuevo usuario")
        btn_crear.setCursor(Qt.PointingHandCursor)
        btn_crear.setStyleSheet("""
            QPushButton {
                background-color: #2563eb; color: white;
                padding: 12px 22px; border: none;
                border-radius: 8px; font-size: 14px; font-weight: bold;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        btn_crear.clicked.connect(self._abrir_crear_usuario)

        btn_ver = QPushButton("👤  Ver usuarios")
        btn_ver.setCursor(Qt.PointingHandCursor)
        btn_ver.setStyleSheet("""
            QPushButton {
                background-color: #f1f5f9; color: #1a2332;
                padding: 12px 22px; border: 1.5px solid #e2e8f0;
                border-radius: 8px; font-size: 14px; font-weight: bold;
            }
            QPushButton:hover { background-color: #e2e8f0; }
        """)
        btn_ver.clicked.connect(self._abrir_ver_usuarios)

        btns_row = QHBoxLayout()
        btns_row.setSpacing(12)
        btns_row.addWidget(btn_crear)
        btns_row.addWidget(btn_ver)
        btns_row.addStretch()
        outer.addLayout(btns_row)

        return card

    def _abrir_crear_usuario(self):
        """Abre el diálogo para crear un usuario."""
        dialog = CrearUsuarioDialog(self)
        dialog.exec()

    def _abrir_ver_usuarios(self):
        """Abre el diálogo para ver y eliminar usuarios."""
        dialog = VerUsuariosDialog(usuario_activo=self._usuario_activo, parent=self)
        dialog.exec()

    def set_usuario(self, username: str, role: str):
        """Muestra u oculta la sección de usuarios según el rol."""
        self._usuario_activo = username
        self._usuario_role   = role
        es_privilegiado = role == 'admin' or username == 'prueba'
        self._grupo_usuarios.setVisible(es_privilegiado)
        # Mostrar botón limpiar historial solo a admin
        if hasattr(self, '_btn_limpiar_hist'):
            self._btn_limpiar_hist.setVisible(es_privilegiado)
        # Populate usuario dropdown and load history
        if hasattr(self, '_cmb_hist_usuario'):
            self._actualizar_usuarios_dropdown()
            self._aplicar_filtros_historial()
        # Refresh resumen tab
        if hasattr(self, '_resumen_cmb_modulo'):
            self._actualizar_resumen()

    def crear_grupo_gimnasio(self):
        """Crea la card de información del gimnasio."""
        _CARD_SS = """
            QFrame#card {
                background-color: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 12px;
            }
        """
        _INPUT_SS = """
            QLineEdit {
                padding: 9px 12px;
                border: 1.5px solid #e2e8f0;
                border-radius: 7px;
                background-color: #f8fafc;
                color: #1a2332;
                font-size: 13px;
            }
            QLineEdit:focus {
                border: 1.5px solid #94a3b8;
                background-color: #ffffff;
            }
        """
        _LBL_SS = "color: #64748b; font-size: 12px; font-weight: bold; background: transparent; border: none;"

        def _lbl(text):
            l = QLabel(text)
            l.setStyleSheet(_LBL_SS)
            return l

        card = QFrame()
        card.setObjectName("card")
        card.setStyleSheet(_CARD_SS)

        outer = QVBoxLayout(card)
        outer.setContentsMargins(24, 20, 24, 24)
        outer.setSpacing(14)

        # ── Cabecera ─────────────────────────────────────────────────────
        hdr = QHBoxLayout()
        ico = QLabel("🏢")
        ico.setFont(QFont("Arial", 16))
        ico.setStyleSheet("background: transparent; border: none;")
        lbl_titulo = QLabel("Información del Gimnasio")
        lbl_titulo.setFont(QFont("Arial", 14, QFont.Bold))
        lbl_titulo.setStyleSheet("color: #1a2332; background: transparent; border: none;")
        hdr.addWidget(ico)
        hdr.addSpacing(6)
        hdr.addWidget(lbl_titulo)
        hdr.addStretch()
        outer.addLayout(hdr)

        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet("background-color: #f0f4f8; border: none; max-height: 1px;")
        outer.addWidget(sep)

        # ── Campos ───────────────────────────────────────────────────────
        self.txt_nombre_gym = QLineEdit()
        self.txt_nombre_gym.setPlaceholderText("Ej: KyoGym")
        self.txt_nombre_gym.setValidator(crear_validador_nombre())
        self.txt_nombre_gym.setStyleSheet(_INPUT_SS)

        self.txt_direccion = QLineEdit()
        self.txt_direccion.setPlaceholderText("Ej: Calle Principal #123")
        self.txt_direccion.setStyleSheet(_INPUT_SS)

        self.txt_telefono = TelefonoFormateadoLineEdit()
        self.txt_telefono.setStyleSheet(_INPUT_SS)

        self.txt_email = QLineEdit()
        self.txt_email.setPlaceholderText("Ej: info@kyogym.com")
        self.txt_email.setValidator(crear_validador_email())
        self.txt_email.setStyleSheet(_INPUT_SS)

        self.txt_rfc = QLineEdit()
        self.txt_rfc.setPlaceholderText("Ej: ABC123456XYZ")
        self.txt_rfc.setStyleSheet(_INPUT_SS)

        # ── Grid 2 columnas ───────────────────────────────────────────────
        grid = QGridLayout()
        grid.setHorizontalSpacing(20)
        grid.setVerticalSpacing(5)
        grid.setColumnStretch(0, 1)
        grid.setColumnStretch(1, 1)

        grid.addWidget(_lbl("Nombre del gimnasio"), 0, 0)
        grid.addWidget(_lbl("Dirección"),           0, 1)
        grid.addWidget(self.txt_nombre_gym,         1, 0)
        grid.addWidget(self.txt_direccion,          1, 1)

        grid.addWidget(_lbl("Teléfono"),  3, 0)
        grid.addWidget(_lbl("Email"),     3, 1)
        grid.addWidget(self.txt_telefono, 4, 0)
        grid.addWidget(self.txt_email,    4, 1)

        grid.addWidget(_lbl("RFC / NIT"), 6, 0, 1, 2)
        grid.addWidget(self.txt_rfc,      7, 0, 1, 2)

        grid.setRowMinimumHeight(2, 8)
        grid.setRowMinimumHeight(5, 8)

        outer.addLayout(grid)
        return card

    def crear_grupo_logo(self):
        """Crea la card de logo del gimnasio."""
        _CARD_SS = """
            QFrame#card {
                background-color: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 12px;
            }
        """
        card = QFrame()
        card.setObjectName("card")
        card.setStyleSheet(_CARD_SS)

        outer = QVBoxLayout(card)
        outer.setContentsMargins(24, 20, 24, 24)
        outer.setSpacing(14)

        # ── Cabecera ─────────────────────────────────────────────────────
        hdr = QHBoxLayout()
        ico = QLabel("🖼️")
        ico.setFont(QFont("Arial", 16))
        ico.setStyleSheet("background: transparent; border: none;")
        lbl_titulo = QLabel("Logo del Gimnasio")
        lbl_titulo.setFont(QFont("Arial", 14, QFont.Bold))
        lbl_titulo.setStyleSheet("color: #1a2332; background: transparent; border: none;")
        hdr.addWidget(ico)
        hdr.addSpacing(6)
        hdr.addWidget(lbl_titulo)
        hdr.addStretch()
        outer.addLayout(hdr)

        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet("background-color: #f0f4f8; border: none; max-height: 1px;")
        outer.addWidget(sep)

        # ── Vista previa centrada ─────────────────────────────────────────
        self.logo_preview = QLabel()
        self.logo_preview.setFixedSize(110, 110)
        self.logo_preview.setAlignment(Qt.AlignCenter)
        self.logo_preview.setStyleSheet("""
            background-color: #f8fafc;
            border: 2px dashed #cbd5e1;
            border-radius: 10px;
            color: #94a3b8;
            font-size: 12px;
        """)
        self.actualizar_preview_logo()

        preview_wrap = QHBoxLayout()
        preview_wrap.addStretch()
        preview_wrap.addWidget(self.logo_preview)
        preview_wrap.addStretch()
        outer.addLayout(preview_wrap)

        # ── Botones ───────────────────────────────────────────────────────
        btn_seleccionar = QPushButton("📁  Seleccionar imagen")
        btn_seleccionar.setCursor(Qt.PointingHandCursor)
        btn_seleccionar.setStyleSheet("""
            QPushButton {
                background-color: #2563eb; color: white;
                padding: 9px 16px; border: none;
                border-radius: 7px; font-size: 13px; font-weight: bold;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        btn_seleccionar.clicked.connect(self.seleccionar_logo)

        btn_eliminar = QPushButton("🗑️  Eliminar logo")
        btn_eliminar.setCursor(Qt.PointingHandCursor)
        btn_eliminar.setStyleSheet("""
            QPushButton {
                background-color: #f1f5f9; color: #64748b;
                padding: 9px 16px; border: 1.5px solid #e2e8f0;
                border-radius: 7px; font-size: 13px; font-weight: bold;
            }
            QPushButton:hover {
                background-color: #fee2e2; color: #dc2626; border-color: #fca5a5;
            }
        """)
        btn_eliminar.clicked.connect(self.eliminar_logo)

        btns_row = QHBoxLayout()
        btns_row.setSpacing(10)
        btns_row.addWidget(btn_seleccionar)
        btns_row.addWidget(btn_eliminar)
        outer.addLayout(btns_row)

        info = QLabel("PNG, JPG, ICO · Tamaño recomendado: 200×200 px")
        info.setAlignment(Qt.AlignCenter)
        info.setStyleSheet("color: #94a3b8; font-size: 11px; background: transparent; border: none;")
        outer.addWidget(info)
        outer.addStretch()

        return card

    def aplicar_estilo_input(self, widget):
        """Aplica estilo a los inputs"""
        widget.setStyleSheet("""
            QLineEdit, QSpinBox, QComboBox {
                padding: 10px;
                border: 2px solid #d0d0d0;
                border-radius: 6px;
                background-color: #f5f5f5;
                color: #1a1a1a;
                font-size: 13px;
            }
            QLineEdit:focus, QSpinBox:focus, QComboBox:focus {
                border: 2px solid #c0c0c0;
            }
            QComboBox QAbstractItemView {
                background-color: #f5f5f5;
                color: #1a1a1a;
                selection-background-color: #808080;
                selection-color: white;
            }
            QComboBox QAbstractItemView::item {
                color: #1a1a1a;
            }
            QCalendarWidget QAbstractItemView {
                selection-background-color: #808080;
                selection-color: white;
                color: #1a1a1a;
                background-color: #f5f5f5;
            }
            QCalendarWidget QWidget {
                color: #1a1a1a;
                background-color: #f5f5f5;
            }
            QCalendarWidget QWidget#qt_calendar_navigationbar {
                background-color: #e8e8e8;
            }
            QCalendarWidget QToolButton {
                color: #555555;
                background-color: #e8e8e8;
            }
            QCalendarWidget QMenu {
                color: #1a1a1a;
                background-color: #f5f5f5;
            }
            QCalendarWidget QSpinBox {
                color: #1a1a1a;
                background-color: #f5f5f5;
            }
            QCalendarWidget QTableView {
                color: #1a1a1a;
                background-color: #f5f5f5;
            }
            QPushButton {
                color: #1a1a1a;
            }
        """)
    
    def actualizar_preview_logo(self):
        """Actualiza el preview del logo"""
        if os.path.exists(self.logo_path):
            pixmap = QPixmap(self.logo_path)
            scaled_pixmap = pixmap.scaled(100, 100, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.logo_preview.setPixmap(scaled_pixmap)
        else:
            self.logo_preview.setText("Sin logo\n📷")
            self.logo_preview.setStyleSheet("""
                border: 2px dashed #c0c0c0;
                background-color: #f5f5f5;
                color: #666666;
                font-size: 12px;
                border-radius: 8px;
            """)
    
    def seleccionar_logo(self):
        """Permite seleccionar un nuevo logo"""
        archivo, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar Logo",
            "",
            "Imágenes (*.png *.jpg *.jpeg *.bmp *.ico)"
        )
        
        if archivo:
            try:
                # Copiar el archivo a la carpeta assets
                import shutil
                destino = self.logo_path
                shutil.copy2(archivo, destino)
                
                self.actualizar_preview_logo()
                # Mensaje con estilo
                msg = QMessageBox(self)
                msg.setWindowTitle("Éxito")
                msg.setText("Logo actualizado correctamente.\nReinicie la aplicación para ver los cambios.")
                msg.setStyleSheet("""
                    QMessageBox {
                        background-color: #f5f5f5;
                    }
                    QLabel {
                        color: #2c2c2c;
                        font-size: 13px;
                        min-width: 300px;
                    }
                    QPushButton {
                        background-color: #27ae60;
                        color: white;
                        padding: 8px 20px;
                        border: none;
                        border-radius: 4px;
                        font-weight: bold;
                        font-size: 13px;
                        min-width: 80px;
                    }
                    QPushButton:hover {
                        background-color: #229954;
                    }
                """)
                msg.exec()
            except Exception as e:
                msg = QMessageBox(self)
                msg.setWindowTitle("Error")
                msg.setText(f"Error al copiar el logo:\n{str(e)}")
                msg.setStyleSheet("""
                    QMessageBox {
                        background-color: #f5f5f5;
                    }
                    QLabel {
                        color: #2c2c2c;
                        font-size: 13px;
                        min-width: 300px;
                    }
                    QPushButton {
                        background-color: #e74c3c;
                        color: white;
                        padding: 8px 20px;
                        border: none;
                        border-radius: 4px;
                        font-weight: bold;
                        font-size: 13px;
                        min-width: 80px;
                    }
                    QPushButton:hover {
                        background-color: #c0392b;
                    }
                """)
                msg.exec()
    
    def eliminar_logo(self):
        """Elimina el logo actual"""
        msg = QMessageBox(self)
        msg.setWindowTitle("Confirmar")
        msg.setText("¿Está seguro de eliminar el logo?")
        msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        msg.setDefaultButton(QMessageBox.No)
        msg.setStyleSheet("""
            QMessageBox {
                background-color: #ffffff;
            }
            QLabel {
                color: #2c2c2c;
                font-size: 13px;
                min-width: 300px;
            }
            QPushButton {
                background-color: #2c3e50;
                color: white;
                padding: 8px 20px;
                border: none;
                border-radius: 4px;
                font-weight: bold;
                font-size: 13px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #3d5166;
            }
        """)
        respuesta = msg.exec()
        
        if respuesta == QMessageBox.Yes:
            try:
                if os.path.exists(self.logo_path):
                    os.remove(self.logo_path)
                    self.actualizar_preview_logo()
                    # Mensaje con estilo
                    msg = QMessageBox(self)
                    msg.setWindowTitle("Éxito")
                    msg.setText("Logo eliminado correctamente.")
                    msg.setStyleSheet("""
                        QMessageBox {
                            background-color: #f5f5f5;
                        }
                        QLabel {
                            color: #2c2c2c;
                            font-size: 13px;
                            min-width: 300px;
                        }
                        QPushButton {
                            background-color: #27ae60;
                            color: white;
                            padding: 8px 20px;
                            border: none;
                            border-radius: 4px;
                            font-weight: bold;
                            font-size: 13px;
                            min-width: 80px;
                        }
                        QPushButton:hover {
                            background-color: #229954;
                        }
                    """)
                    msg.exec()
            except Exception as e:
                msg = QMessageBox(self)
                msg.setWindowTitle("Error")
                msg.setText(f"Error al eliminar el logo:\n{str(e)}")
                msg.setStyleSheet("""
                    QMessageBox {
                        background-color: #f5f5f5;
                    }
                    QLabel {
                        color: #2c2c2c;
                        font-size: 13px;
                        min-width: 300px;
                    }
                    QPushButton {
                        background-color: #e74c3c;
                        color: white;
                        padding: 8px 20px;
                        border: none;
                        border-radius: 4px;
                        font-weight: bold;
                        font-size: 13px;
                        min-width: 80px;
                    }
                    QPushButton:hover {
                        background-color: #c0392b;
                    }
                """)
                msg.exec()
    
    def cargar_configuracion(self):
        """Carga la configuración desde el archivo JSON"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    
                    # Información del gimnasio
                    self.txt_nombre_gym.setText(config.get('nombre_gimnasio', 'KyoGym'))
                    self.txt_direccion.setText(config.get('direccion', ''))
                    self.txt_telefono.setText(config.get('telefono', ''))
                    self.txt_email.setText(config.get('email', ''))
                    self.txt_rfc.setText(config.get('rfc', ''))
                    
        except Exception as e:
            msg = QMessageBox(self)
            msg.setWindowTitle("Advertencia")
            msg.setText(f"Error al cargar configuración:\n{str(e)}\n\nSe usarán valores por defecto.")
            msg.setStyleSheet("""
                QMessageBox {
                    background-color: #f5f5f5;
                }
                QLabel {
                    color: #2c2c2c;
                    font-size: 13px;
                    min-width: 300px;
                }
                QPushButton {
                    background-color: #f39c12;
                    color: white;
                    padding: 8px 20px;
                    border: none;
                    border-radius: 4px;
                    font-weight: bold;
                    font-size: 13px;
                    min-width: 80px;
                }
                QPushButton:hover {
                    background-color: #e67e22;
                }
            """)
            msg.exec()
    
    def guardar_configuracion(self):
        """Guarda la configuración en el archivo JSON"""
        try:
            dias_alerta = DEFAULT_DIAS_ALERTA_VENCIMIENTO
            if os.path.exists(self.config_file):
                try:
                    with open(self.config_file, 'r', encoding='utf-8') as f:
                        config_existente = json.load(f)
                        dias_alerta = int(config_existente.get('dias_alerta_vencimiento', DEFAULT_DIAS_ALERTA_VENCIMIENTO))
                        if dias_alerta < 0:
                            dias_alerta = DEFAULT_DIAS_ALERTA_VENCIMIENTO
                except (ValueError, TypeError, json.JSONDecodeError):
                    dias_alerta = DEFAULT_DIAS_ALERTA_VENCIMIENTO

            config = {
                'nombre_gimnasio': self.txt_nombre_gym.text(),
                'direccion': self.txt_direccion.text(),
                'telefono': self.txt_telefono.text(),
                'email': self.txt_email.text(),
                'rfc': self.txt_rfc.text(),
                'dias_alerta_vencimiento': dias_alerta
            }
            
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=4, ensure_ascii=False)
            
            # Mensaje con estilo
            msg = QMessageBox(self)
            msg.setWindowTitle("Éxito")
            msg.setText("Configuración guardada correctamente.")
            msg.setStyleSheet("""
                QMessageBox {
                    background-color: #f5f5f5;
                }
                QLabel {
                    color: #2c2c2c;
                    font-size: 13px;
                    min-width: 300px;
                }
                QPushButton {
                    background-color: #27ae60;
                    color: white;
                    padding: 8px 20px;
                    border: none;
                    border-radius: 4px;
                    font-weight: bold;
                    font-size: 13px;
                    min-width: 80px;
                }
                QPushButton:hover {
                    background-color: #229954;
                }
            """)
            msg.exec()
        except Exception as e:
            msg = QMessageBox(self)
            msg.setWindowTitle("Error")
            msg.setText(f"Error al guardar configuración:\n{str(e)}")
            msg.setStyleSheet("""
                QMessageBox {
                    background-color: #f5f5f5;
                }
                QLabel {
                    color: #2c2c2c;
                    font-size: 13px;
                    min-width: 300px;
                }
                QPushButton {
                    background-color: #e74c3c;
                    color: white;
                    padding: 8px 20px;
                    border: none;
                    border-radius: 4px;
                    font-weight: bold;
                    font-size: 13px;
                    min-width: 80px;
                }
                QPushButton:hover {
                    background-color: #c0392b;
                }
            """)
            msg.exec()
    
    def restaurar_predeterminados(self):
        """Restaura los valores predeterminados"""
        msg = QMessageBox(self)
        msg.setWindowTitle("Confirmar")
        msg.setText("¿Está seguro de restaurar los valores predeterminados?")
        msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        msg.setDefaultButton(QMessageBox.No)
        msg.setStyleSheet("""
            QMessageBox {
                background-color: #ffffff;
            }
            QLabel {
                color: #2c2c2c;
                font-size: 13px;
                min-width: 300px;
            }
            QPushButton {
                background-color: #2c3e50;
                color: white;
                padding: 8px 20px;
                border: none;
                border-radius: 4px;
                font-weight: bold;
                font-size: 13px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #3d5166;
            }
        """)
        respuesta = msg.exec()
        
        if respuesta == QMessageBox.Yes:
            self.txt_nombre_gym.setText("KyoGym")
            self.txt_direccion.setText("")
            self.txt_telefono.setText("")
            self.txt_email.setText("")
            # Restablecer demás campos
            self.txt_rfc.setText("")

    # -----------------------------------------------------------------------
    # Pestaña Historial de Acciones
    # -----------------------------------------------------------------------

    # Colores de fondo por módulo
    _COLORES_MODULO = {
        "Clientes":   "#eaf4ff",
        "Pagos":      "#eafff2",
        "Membresías": "#fffbe6",
        "Inventario": "#f5eeff",
    }
    _COLORES_TEXTO_MOD = {
        "Clientes":   "#1558a8",
        "Pagos":      "#1a6b3a",
        "Membresías": "#8a5c00",
        "Inventario": "#5c2d91",
    }

    # Colores de badge por acción
    _BADGE_COLORES = {
        "CREAR":     ("#1abc9c", "#ffffff"),
        "MODIFICAR": ("#f39c12", "#ffffff"),
        "EDITAR":    ("#f39c12", "#ffffff"),
        "ELIMINAR":  ("#e74c3c", "#ffffff"),
        "PAGO":      ("#e67e22", "#ffffff"),
        "IMPORTAR":  ("#9b59b6", "#ffffff"),
        "RENOVAR":   ("#2980b9", "#ffffff"),
    }
    _BADGE_DEFAULT = ("#7f8c8d", "#ffffff")

    def _crear_tab_historial(self):
        """Crea la pestaña completa de historial de acciones."""
        tab = QWidget()
        tab.setStyleSheet("background-color: #f5f7fa;")
        outer = QVBoxLayout(tab)
        outer.setContentsMargins(24, 24, 24, 16)
        outer.setSpacing(16)

        # ── Cabecera de sección ─────────────────────────────────────────
        cab_layout = QHBoxLayout()
        lbl_titulo = QLabel("📋  Historial de Acciones")
        lbl_titulo.setFont(QFont("Arial", 17, QFont.Bold))
        lbl_titulo.setStyleSheet("color: #1a1a1a; background: transparent; border: none;")
        cab_layout.addWidget(lbl_titulo)
        cab_layout.addStretch()

        self._btn_limpiar_hist = QPushButton("🗑️  Limpiar historial")
        self._btn_limpiar_hist.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c; color: white;
                padding: 8px 18px; border: none; border-radius: 6px;
                font-weight: bold; font-size: 13px;
            }
            QPushButton:hover { background-color: #c0392b; }
        """)
        self._btn_limpiar_hist.setVisible(False)
        self._btn_limpiar_hist.clicked.connect(self._limpiar_historial)
        cab_layout.addWidget(self._btn_limpiar_hist)
        outer.addLayout(cab_layout)

        lbl_sub = QLabel("Registro completo de todas las acciones realizadas en el sistema, seleccionando.")
        lbl_sub.setStyleSheet("color: #888888; font-size: 13px; background: transparent; border: none;")
        outer.addWidget(lbl_sub)

        # ── Tarjeta de filtros ──────────────────────────────────────────
        card_filtros = QFrame()
        card_filtros.setStyleSheet("""
            QFrame {
                background-color: #ffffff;
                border: 1px solid #e4e8ef;
                border-radius: 10px;
            }
        """)
        filtros_outer = QVBoxLayout(card_filtros)
        filtros_outer.setContentsMargins(18, 14, 18, 14)
        filtros_outer.setSpacing(10)

        # Fila de filtros
        fila = QHBoxLayout()
        fila.setSpacing(10)

        _cmb_style = """
            QComboBox {
                padding: 7px 10px; border: 1px solid #d0d0d0;
                border-radius: 6px; background-color: #f9f9f9;
                color: #1a1a1a; font-size: 13px; min-width: 130px;
            }
            QComboBox:focus { border: 1px solid #2c3e50; }
            QComboBox::drop-down { border: none; }
            QComboBox QAbstractItemView {
                background-color: #ffffff; color: #1a1a1a;
                selection-background-color: #2c3e50; selection-color: white;
                border: 1px solid #d0d0d0; border-radius: 4px;
            }
        """
        _input_style = """
            QLineEdit, QDateEdit {
                padding: 7px 10px; border: 1px solid #d0d0d0;
                border-radius: 6px; background-color: #f9f9f9;
                color: #1a1a1a; font-size: 13px;
            }
            QLineEdit:focus, QDateEdit:focus { border: 1px solid #2c3e50; }
        """

        # Label Módulo
        lbl_m = QLabel("Módulo:")
        lbl_m.setStyleSheet("color: #555; font-weight: bold; font-size: 13px; background: transparent; border: none;")
        fila.addWidget(lbl_m)

        self._cmb_hist_modulo = QComboBox()
        self._cmb_hist_modulo.addItems(["Todos", "Clientes", "Pagos", "Membresías", "Inventario"])
        self._cmb_hist_modulo.setStyleSheet(_cmb_style)
        fila.addWidget(self._cmb_hist_modulo)

        # Label Usuario
        lbl_u = QLabel("Usuario:")
        lbl_u.setStyleSheet("color: #555; font-weight: bold; font-size: 13px; background: transparent; border: none;")
        fila.addWidget(lbl_u)

        self._cmb_hist_usuario = QComboBox()
        self._cmb_hist_usuario.addItem("Todos")
        self._cmb_hist_usuario.setStyleSheet(_cmb_style)
        self._actualizar_usuarios_dropdown()
        fila.addWidget(self._cmb_hist_usuario)

        # ── Rango de fechas Desde / Hasta ───────────────────────────────
        _CAL_SS = """
            QCalendarWidget QAbstractItemView {
                selection-background-color: #5e88b4;
                selection-color: white;
                color: #2c2c2c;
                background-color: #eaf0f9;
            }
            QCalendarWidget QWidget {
                color: #2c2c2c;
                background-color: #eaf0f9;
            }
            QCalendarWidget QWidget#qt_calendar_navigationbar {
                background-color: #dce7f3;
            }
            QCalendarWidget QToolButton {
                color: #2c2c2c;
                background-color: #dce7f3;
            }
            QCalendarWidget QMenu {
                color: #2c2c2c;
                background-color: #f5f5f5;
            }
            QCalendarWidget QSpinBox {
                color: #2c2c2c;
                background-color: #f0f0f0;
            }
            QCalendarWidget QTableView {
                color: #2c2c2c;
            }
        """

        lbl_desde = QLabel("Desde:")
        lbl_desde.setStyleSheet("color: #555; font-weight: bold; font-size: 13px; background: transparent; border: none;")
        fila.addWidget(lbl_desde)

        _hoy = QDate.currentDate()
        _primer_dia_mes = QDate(_hoy.year(), _hoy.month(), 1)

        self._date_desde = QDateEdit()
        self._date_desde.setCalendarPopup(True)
        self._date_desde.setDisplayFormat("dd/MM/yyyy")
        self._date_desde.setMinimumDate(QDate(2000, 1, 1))
        self._date_desde.setDate(_primer_dia_mes)
        self._date_desde.setStyleSheet(_input_style + "QDateEdit { min-width: 120px; }")
        self._date_desde.calendarWidget().setStyleSheet(_CAL_SS)
        fila.addWidget(self._date_desde)

        lbl_hasta = QLabel("Hasta:")
        lbl_hasta.setStyleSheet("color: #555; font-weight: bold; font-size: 13px; background: transparent; border: none;")
        fila.addWidget(lbl_hasta)

        self._date_hasta = QDateEdit()
        self._date_hasta.setCalendarPopup(True)
        self._date_hasta.setDisplayFormat("dd/MM/yyyy")
        self._date_hasta.setMinimumDate(QDate(2000, 1, 1))
        self._date_hasta.setDate(_hoy)
        self._date_hasta.setStyleSheet(_input_style + "QDateEdit { min-width: 120px; }")
        self._date_hasta.calendarWidget().setStyleSheet(_CAL_SS)
        fila.addWidget(self._date_hasta)

        # ── Acceso rápido: Hoy ───────────────────────────────────────────
        def _set_hoy():
            hoy = QDate.currentDate()
            self._date_desde.setDate(hoy)
            self._date_hasta.setDate(hoy)
            self._aplicar_filtros_historial()

        b_hoy = QPushButton("Hoy")
        b_hoy.setStyleSheet("""
            QPushButton {
                background-color: #f1f5f9; color: #475569;
                padding: 5px 11px; border: 1.5px solid #e2e8f0;
                border-radius: 12px; font-size: 12px; font-weight: bold;
            }
            QPushButton:hover { background-color: #e2e8f0; color: #1a2332; }
        """)
        b_hoy.setCursor(Qt.PointingHandCursor)
        b_hoy.clicked.connect(_set_hoy)
        fila.addWidget(b_hoy)

        # Buscador
        self._txt_hist_buscar = QLineEdit()
        self._txt_hist_buscar.setPlaceholderText("🔍  Buscar...")
        self._txt_hist_buscar.setStyleSheet(_input_style + "QLineEdit { min-width: 180px; }")
        self._txt_hist_buscar.returnPressed.connect(self._aplicar_filtros_historial)
        fila.addWidget(self._txt_hist_buscar)

        # Botón Actualizar
        btn_act = QPushButton("🔄  Actualizar")
        btn_act.setStyleSheet("""
            QPushButton {
                background-color: #2c3e50; color: white;
                padding: 8px 18px; border: none; border-radius: 6px;
                font-weight: bold; font-size: 13px;
            }
            QPushButton:hover { background-color: #3d5166; }
        """)
        btn_act.clicked.connect(self._aplicar_filtros_historial)
        fila.addWidget(btn_act)

        # Botón Exportar
        btn_exp = QPushButton("📤  Exportar")
        btn_exp.setStyleSheet("""
            QPushButton {
                background-color: #27ae60; color: white;
                padding: 8px 18px; border: none; border-radius: 6px;
                font-weight: bold; font-size: 13px;
            }
            QPushButton:hover { background-color: #229954; }
        """)
        btn_exp.clicked.connect(self._exportar_historial)
        fila.addWidget(btn_exp)

        # Botón Limpiar filtros
        btn_clf = QPushButton("✕  Limpiar filtros")
        btn_clf.setStyleSheet("""
            QPushButton {
                background-color: transparent; color: #888888;
                padding: 8px 12px; border: 1px solid #d0d0d0; border-radius: 6px;
                font-size: 13px;
            }
            QPushButton:hover { background-color: #f0f0f0; color: #333; }
        """)
        btn_clf.clicked.connect(self._limpiar_filtros_historial)
        fila.addWidget(btn_clf)

        fila.addStretch()
        filtros_outer.addLayout(fila)

        # Leyenda
        leyenda = QHBoxLayout()
        leyenda.setSpacing(20)
        for mod, color in self._COLORES_TEXTO_MOD.items():
            dot = QLabel(f"■  {mod}")
            dot.setStyleSheet(
                f"color: {color}; font-size: 12px; font-weight: bold;"
                " background: transparent; border: none;"
            )
            leyenda.addWidget(dot)
        leyenda.addStretch()
        filtros_outer.addLayout(leyenda)

        outer.addWidget(card_filtros)

        # ── Tarjeta de tabla ────────────────────────────────────────────
        card_tabla = QFrame()
        card_tabla.setStyleSheet("""
            QFrame {
                background-color: #ffffff;
                border: 1px solid #e4e8ef;
                border-radius: 10px;
            }
        """)
        tabla_outer = QVBoxLayout(card_tabla)
        tabla_outer.setContentsMargins(0, 0, 0, 0)
        tabla_outer.setSpacing(0)

        self._tabla_hist = QTableWidget()
        self._tabla_hist.setColumnCount(5)
        self._tabla_hist.setHorizontalHeaderLabels([
            "Fecha y Hora", "Módulo", "Acción", "Descripción", "Usuario"
        ])
        self._tabla_hist.setStyleSheet("""
            QTableWidget {
                border: none;
                background-color: #ffffff;
                gridline-color: #f0f0f0;
                font-size: 13px;
                color: #1a1a1a;
                outline: none;
            }
            QTableWidget::item {
                padding: 10px 12px;
                border-bottom: 1px solid #f4f4f4;
            }
            QTableWidget::item:hover {
                background-color: #f0f7ff;
            }
            QHeaderView::section {
                background-color: #2c3e50;
                color: white;
                font-weight: bold;
                font-size: 13px;
                padding: 10px 12px;
                border: none;
                border-right: 1px solid #3d5166;
            }
            QHeaderView::section:last { border-right: none; }
            QScrollBar:vertical {
                background: #f5f5f5; width: 8px; border-radius: 4px;
            }
            QScrollBar::handle:vertical {
                background: #c0c0c0; border-radius: 4px;
            }
        """)
        hh = self._tabla_hist.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(2, QHeaderView.Fixed)
        self._tabla_hist.setColumnWidth(2, 110)
        hh.setSectionResizeMode(3, QHeaderView.Stretch)
        hh.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self._tabla_hist.setEditTriggers(QTableWidget.NoEditTriggers)
        self._tabla_hist.setSelectionBehavior(QTableWidget.SelectRows)
        self._tabla_hist.setSelectionMode(QTableWidget.SingleSelection)
        self._tabla_hist.verticalHeader().setVisible(False)
        self._tabla_hist.setSortingEnabled(False)
        self._tabla_hist.setShowGrid(True)
        self._tabla_hist.setAlternatingRowColors(False)
        tabla_outer.addWidget(self._tabla_hist)

        # Mensaje vacío
        self._lbl_hist_vacio = QLabel("No hay registros de actividad para los filtros seleccionados.")
        self._lbl_hist_vacio.setAlignment(Qt.AlignCenter)
        self._lbl_hist_vacio.setStyleSheet(
            "color: #aaaaaa; font-size: 14px; padding: 40px; background: transparent; border: none;"
        )
        self._lbl_hist_vacio.setVisible(False)
        tabla_outer.addWidget(self._lbl_hist_vacio)

        outer.addWidget(card_tabla, 1)

        # ── Barra de paginación ─────────────────────────────────────────
        pag_card = QFrame()
        pag_card.setStyleSheet("""
            QFrame {
                background-color: #ffffff;
                border: 1px solid #e4e8ef;
                border-radius: 8px;
            }
        """)
        pag_outer = QHBoxLayout(pag_card)
        pag_outer.setContentsMargins(16, 8, 16, 8)
        pag_outer.setSpacing(8)

        # Selector de entradas por página
        self._cmb_por_pagina = QComboBox()
        self._cmb_por_pagina.addItems(["10", "25", "50", "100"])
        self._cmb_por_pagina.setFixedWidth(65)
        self._cmb_por_pagina.setStyleSheet("""
            QComboBox {
                padding: 4px 8px; border: 1px solid #d0d0d0;
                border-radius: 5px; background: #f9f9f9; color: #1a1a1a; font-size: 13px;
            }
            QComboBox QAbstractItemView {
                background: #fff; color: #1a1a1a;
                selection-background-color: #2c3e50; selection-color: white;
            }
        """)
        self._cmb_por_pagina.currentTextChanged.connect(self._cambiar_por_pagina)
        pag_outer.addWidget(self._cmb_por_pagina)

        lbl_ent = QLabel("entradas")
        lbl_ent.setStyleSheet("color: #888; font-size: 13px; background: transparent; border: none;")
        pag_outer.addWidget(lbl_ent)

        pag_outer.addStretch()

        self._lbl_pag_info = QLabel("0 - 0 de 0")
        self._lbl_pag_info.setStyleSheet(
            "color: #888888; font-size: 13px; background: transparent; border: none;"
        )
        pag_outer.addWidget(self._lbl_pag_info)

        # Botones de navegación
        _btn_nav_style = """
            QPushButton {
                background-color: transparent; color: #2c3e50;
                padding: 5px 11px; border: 1px solid #d0d0d0; border-radius: 5px;
                font-size: 13px; font-weight: bold; min-width: 32px;
            }
            QPushButton:hover { background-color: #eef2f7; }
            QPushButton:disabled { color: #cccccc; border-color: #eeeeee; }
            QPushButton[pagina_activa="true"] {
                background-color: #2c3e50; color: white; border-color: #2c3e50;
            }
        """
        self._btn_prev = QPushButton("‹ anteriores")
        self._btn_prev.setStyleSheet(_btn_nav_style)
        self._btn_prev.clicked.connect(self._pagina_anterior)
        pag_outer.addWidget(self._btn_prev)

        # Contenedor de botones de número de página
        self._pag_nums_layout = QHBoxLayout()
        self._pag_nums_layout.setSpacing(4)
        pag_outer.addLayout(self._pag_nums_layout)

        self._btn_next = QPushButton("siguientes ›")
        self._btn_next.setStyleSheet(_btn_nav_style)
        self._btn_next.clicked.connect(self._pagina_siguiente)
        pag_outer.addWidget(self._btn_next)

        outer.addWidget(pag_card)

        # Carga inicial
        self._aplicar_filtros_historial()

        # Conexiones de filtros automáticos
        self._cmb_hist_modulo.currentIndexChanged.connect(self._aplicar_filtros_historial)
        self._cmb_hist_usuario.currentIndexChanged.connect(self._aplicar_filtros_historial)
        self._date_desde.dateChanged.connect(self._aplicar_filtros_historial)
        self._date_hasta.dateChanged.connect(self._aplicar_filtros_historial)

        return tab

    # ── Helpers de historial ────────────────────────────────────────────

    def _actualizar_usuarios_dropdown(self):
        """Rellena el combo de usuarios con todos los existentes en el sistema."""
        try:
            usuarios = get_all_users()
            self._cmb_hist_usuario.blockSignals(True)
            self._cmb_hist_usuario.clear()
            self._cmb_hist_usuario.addItem("Todos")
            for u in usuarios:
                self._cmb_hist_usuario.addItem(u.get('username', ''))
            self._cmb_hist_usuario.blockSignals(False)
        except Exception:
            pass

    def _aplicar_filtros_historial(self):
        """Obtiene todos los registros filtrados y resetea a página 0."""
        modulo  = self._cmb_hist_modulo.currentText()
        usuario = self._cmb_hist_usuario.currentText()
        buscar  = self._txt_hist_buscar.text().strip()

        desde_str = self._date_desde.date().toString("yyyy-MM-dd")
        hasta_str = self._date_hasta.date().toString("yyyy-MM-dd")

        registros = auditoria_service.obtener_historial(
            modulo=modulo  if modulo  != "Todos" else None,
            buscar=buscar  if buscar              else None,
            limite=10000,
        )

        # Filtro por usuario
        if usuario != "Todos":
            registros = [r for r in registros if r.get("usuario") == usuario]

        # Filtro por rango de fecha
        registros = [
            r for r in registros
            if desde_str <= r.get("fecha_hora", "")[:10] <= hasta_str
        ]

        self._hist_todos  = registros
        self._hist_pagina = 0
        self._actualizar_tabla_pagina()
        self._actualizar_paginator()

    def _actualizar_tabla_pagina(self):
        """Renderiza la página actual de la tabla."""
        por_pagina = self._hist_por_pagina
        inicio     = self._hist_pagina * por_pagina
        fin        = inicio + por_pagina
        pagina     = self._hist_todos[inicio:fin]

        limpiar_tabla(self._tabla_hist)

        if not self._hist_todos:
            self._tabla_hist.setVisible(False)
            self._lbl_hist_vacio.setVisible(True)
            return

        self._tabla_hist.setVisible(True)
        self._lbl_hist_vacio.setVisible(False)
        self._tabla_hist.setRowCount(len(pagina))

        for i, reg in enumerate(pagina):
            mod    = reg.get("modulo", "")
            accion = reg.get("accion", "")
            color_fondo = QColor(self._COLORES_MODULO.get(mod, "#ffffff"))

            self._tabla_hist.setRowHeight(i, 44)

            # Col 0 – Fecha
            item0 = QTableWidgetItem(reg.get("fecha_hora", ""))
            item0.setBackground(QColor("#ffffff"))
            item0.setForeground(QColor("#666666"))
            self._tabla_hist.setItem(i, 0, item0)

            # Col 1 – Módulo (con color propio)
            item1 = QTableWidgetItem(mod)
            item1.setBackground(color_fondo)
            item1.setForeground(QColor(self._COLORES_TEXTO_MOD.get(mod, "#2c2c2c")))
            f = item1.font()
            f.setBold(True)
            item1.setFont(f)
            self._tabla_hist.setItem(i, 1, item1)

            # Col 2 – Acción (texto coloreado en negrita)
            item2 = QTableWidgetItem(accion)
            item2.setBackground(QColor("#ffffff"))
            color_acc = self._BADGE_COLORES.get(accion, self._BADGE_DEFAULT)[0]
            item2.setForeground(QColor(color_acc))
            f2 = item2.font()
            f2.setBold(True)
            item2.setFont(f2)
            self._tabla_hist.setItem(i, 2, item2)

            # Col 3 – Descripción
            desc = reg.get("descripcion", "")
            item3 = QTableWidgetItem(desc)
            item3.setBackground(QColor("#ffffff"))
            item3.setToolTip(desc)
            self._tabla_hist.setItem(i, 3, item3)

            # Col 4 – Usuario
            item4 = QTableWidgetItem(reg.get("usuario", ""))
            item4.setBackground(QColor("#ffffff"))
            item4.setForeground(QColor("#555555"))
            self._tabla_hist.setItem(i, 4, item4)

    def _actualizar_paginator(self):
        """Refresca la barra de paginación."""
        total      = len(self._hist_todos)
        por_pagina = self._hist_por_pagina
        paginas    = max(1, (total + por_pagina - 1) // por_pagina)
        inicio     = self._hist_pagina * por_pagina + 1 if total else 0
        fin        = min((self._hist_pagina + 1) * por_pagina, total)

        self._lbl_pag_info.setText(f"{inicio} - {fin} de {total}")
        self._btn_prev.setEnabled(self._hist_pagina > 0)
        self._btn_next.setEnabled(self._hist_pagina < paginas - 1)

        # Limpiar botones numéricos
        while self._pag_nums_layout.count():
            item = self._pag_nums_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        # Mostrar máx 7 números de página con elipsis
        _btn_pag_style = """
            QPushButton {
                background-color: transparent; color: #2c3e50;
                padding: 5px 10px; border: 1px solid #d0d0d0;
                border-radius: 5px; font-size: 13px; min-width: 32px;
            }
            QPushButton:hover { background-color: #eef2f7; }
        """
        _btn_pag_activo = """
            QPushButton {
                background-color: #2c3e50; color: white;
                padding: 5px 10px; border: 1px solid #2c3e50;
                border-radius: 5px; font-size: 13px; min-width: 32px;
                font-weight: bold;
            }
        """

        visible = self._paginas_visibles(self._hist_pagina, paginas)
        prev_n  = None
        for n in visible:
            if prev_n is not None and n - prev_n > 1:
                sep = QLabel("…")
                sep.setStyleSheet(
                    "color: #aaa; font-size: 13px; padding: 0 4px;"
                    " background: transparent; border: none;"
                )
                self._pag_nums_layout.addWidget(sep)
            btn = QPushButton(str(n + 1))
            btn.setStyleSheet(_btn_pag_activo if n == self._hist_pagina else _btn_pag_style)
            btn.setFixedWidth(36)
            btn.clicked.connect(lambda checked, p=n: self._ir_pagina(p))
            self._pag_nums_layout.addWidget(btn)
            prev_n = n

    @staticmethod
    def _paginas_visibles(actual, total, ventana=5):
        """Devuelve lista de índices de página visibles centrados en 'actual'."""
        if total <= ventana + 2:
            return list(range(total))
        mitad  = ventana // 2
        inicio = max(0, min(actual - mitad, total - ventana))
        fin    = min(inicio + ventana, total)
        paginas = list(range(inicio, fin))
        if 0 not in paginas:
            paginas = [0] + paginas
        if total - 1 not in paginas:
            paginas = paginas + [total - 1]
        return paginas

    def _ir_pagina(self, n):
        self._hist_pagina = n
        self._actualizar_tabla_pagina()
        self._actualizar_paginator()

    def _pagina_anterior(self):
        if self._hist_pagina > 0:
            self._ir_pagina(self._hist_pagina - 1)

    def _pagina_siguiente(self):
        paginas = max(1, (len(self._hist_todos) + self._hist_por_pagina - 1) // self._hist_por_pagina)
        if self._hist_pagina < paginas - 1:
            self._ir_pagina(self._hist_pagina + 1)

    def _cambiar_por_pagina(self, texto):
        try:
            self._hist_por_pagina = int(texto)
        except ValueError:
            self._hist_por_pagina = 10
        self._hist_pagina = 0
        self._actualizar_tabla_pagina()
        self._actualizar_paginator()

    def _limpiar_filtros_historial(self):
        """Resetea todos los filtros a su valor predeterminado."""
        self._cmb_hist_modulo.setCurrentIndex(0)
        self._cmb_hist_usuario.setCurrentIndex(0)
        hoy = QDate.currentDate()
        self._date_desde.setDate(QDate(hoy.year(), hoy.month(), 1))
        self._date_hasta.setDate(hoy)
        self._txt_hist_buscar.clear()
        self._aplicar_filtros_historial()

    def _exportar_historial(self):
        """Muestra opciones para exportar el historial (Excel o PDF)."""
        if not self._hist_todos:
            _m = QMessageBox(self); _m.setWindowTitle("Sin datos"); _m.setText("No hay registros para exportar."); _m.setStyleSheet(_MSG_SS); _m.exec()
            return
        filepath, filtro = QFileDialog.getSaveFileName(
            self, "Exportar historial", "historial_acciones",
            "Excel (*.xlsx);;PDF (*.pdf)"
        )
        if not filepath:
            return
        try:
            if filtro.startswith("Excel") or filepath.endswith(".xlsx"):
                if not filepath.endswith(".xlsx"):
                    filepath += ".xlsx"
                self._exportar_excel(filepath)
            else:
                if not filepath.endswith(".pdf"):
                    filepath += ".pdf"
                self._exportar_pdf(filepath)
            _m = QMessageBox(self); _m.setWindowTitle("Exportación completada"); _m.setText(f"Historial exportado correctamente:\n{filepath}"); _m.setStyleSheet(_MSG_SS); _m.exec()
        except ImportError as e:
            _m = QMessageBox(self); _m.setWindowTitle("Dependencia faltante"); _m.setText(str(e)); _m.setStyleSheet(_MSG_SS); _m.exec()
        except Exception as e:
            _m = QMessageBox(self); _m.setWindowTitle("Error al exportar"); _m.setText(str(e)); _m.setStyleSheet(_MSG_SS); _m.exec()

    def _exportar_excel(self, filepath):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl no está instalado. Instálalo con: pip install openpyxl")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historial"

        headers_excel = ["Fecha y Hora", "Módulo", "Acción", "Descripción", "Usuario"]
        header_fill = PatternFill("solid", fgColor="2C3E50")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        for col, h in enumerate(headers_excel, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill  = header_fill
            cell.font  = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        col_widths = [20, 14, 14, 60, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        for row_idx, reg in enumerate(self._hist_todos, 2):
            vals = [
                reg.get("fecha_hora",   ""),
                reg.get("modulo",       ""),
                reg.get("accion",       ""),
                reg.get("descripcion",  ""),
                reg.get("usuario",      ""),
            ]
            for col_idx, val in enumerate(vals, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        # ── Hoja Empleados ──────────────────────────────────────────────
        ws_emp = wb.create_sheet("Empleados")

        periodo_desde = self._date_desde.date().toString("dd/MM/yyyy")
        periodo_hasta = self._date_hasta.date().toString("dd/MM/yyyy")

        # Fila de título con el período
        title_fill = PatternFill("solid", fgColor="2C3E50")
        title_font = Font(bold=True, color="FFFFFF", size=13)
        ws_emp.merge_cells("A1:H1")
        title_cell = ws_emp["A1"]
        title_cell.value = f"Resumen de actividad por empleado  |  Período: {periodo_desde} – {periodo_hasta}"
        title_cell.fill  = title_fill
        title_cell.font  = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_emp.row_dimensions[1].height = 22

        # Cabeceras
        emp_headers = ["Empleado", "Total acciones", "Clientes", "Pagos", "Membresías", "Inventario", "Otros", "% del total"]
        sub_fill = PatternFill("solid", fgColor="3D5166")
        sub_font = Font(bold=True, color="FFFFFF", size=11)
        for col_i, h in enumerate(emp_headers, 1):
            cell = ws_emp.cell(row=2, column=col_i, value=h)
            cell.fill = sub_fill
            cell.font = sub_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        _MODULOS_CONOCIDOS = ["Clientes", "Pagos", "Membresías", "Inventario"]

        # Agregar por usuario
        from collections import defaultdict
        emp_data: dict[str, dict] = defaultdict(lambda: defaultdict(int))
        for reg in self._hist_todos:
            usr = reg.get("usuario", "—") or "—"
            mod = reg.get("modulo", "Otro") or "Otro"
            emp_data[usr][mod] += 1

        total_global = sum(sum(d.values()) for d in emp_data.values())

        # Colores alternos para filas
        fill_par   = PatternFill("solid", fgColor="EAF0F9")
        fill_impar = PatternFill("solid", fgColor="FFFFFF")
        font_normal = Font(size=11)
        font_bold   = Font(bold=True, size=11)

        emp_sorted = sorted(emp_data.items(), key=lambda x: -sum(x[1].values()))
        for row_i, (usr, mods) in enumerate(emp_sorted, 3):
            total_usr  = sum(mods.values())
            clientes   = mods.get("Clientes",  0)
            pagos      = mods.get("Pagos",     0)
            membresias = mods.get("Membresías",0)
            inventario = mods.get("Inventario",0)
            otros      = sum(v for k, v in mods.items() if k not in _MODULOS_CONOCIDOS)
            pct        = f"{(total_usr / total_global * 100):.1f}%" if total_global else "0%"

            fila_fill = fill_par if row_i % 2 == 0 else fill_impar
            vals_emp = [usr, total_usr, clientes, pagos, membresias, inventario, otros, pct]
            for col_i, val in enumerate(vals_emp, 1):
                cell = ws_emp.cell(row=row_i, column=col_i, value=val)
                cell.fill = fila_fill
                cell.font = font_bold if col_i == 1 else font_normal
                cell.alignment = Alignment(
                    horizontal="left" if col_i == 1 else "center",
                    vertical="center"
                )

        # Anchos de columna
        emp_col_widths = [24, 16, 12, 12, 14, 14, 10, 14]
        for i, w in enumerate(emp_col_widths, 1):
            ws_emp.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        wb.save(filepath)

    def _exportar_pdf(self, filepath):
        try:
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
        except ImportError:
            raise ImportError("reportlab no está instalado. Instálalo con: pip install reportlab")

        doc = SimpleDocTemplate(filepath, pagesize=landscape(A4),
                                leftMargin=20, rightMargin=20, topMargin=30, bottomMargin=30)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Historial de Acciones — KyoGym", styles["Title"]))
        elements.append(Spacer(1, 12))

        data = [["Fecha y Hora", "Módulo", "Acción", "Descripción", "Usuario"]]
        for reg in self._hist_todos:
            data.append([
                reg.get("fecha_hora",   ""),
                reg.get("modulo",       ""),
                reg.get("accion",       ""),
                reg.get("descripcion",  ""),
                reg.get("usuario",      ""),
            ])

        table = Table(data, colWidths=[110, 70, 65, 280, 65], repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#2C3E50")),
            ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
            ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, 0), 10),
            ("ALIGN",       (0, 0), (-1, -1), "LEFT"),
            ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
            ("FONTSIZE",    (0, 1), (-1, -1), 8),
            ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#e0e0e0")),
            ("LEFTPADDING",  (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING",   (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
        ]))
        elements.append(table)
        doc.build(elements)

    def _limpiar_historial(self):
        """Limpia el historial global (solo admin)."""
        msg = QMessageBox(self)
        msg.setWindowTitle("Confirmar")
        msg.setText(
            "¿Eliminar todo el historial de acciones?\n"
            "Esta operación no se puede deshacer.\n\n"
            "Nota: el historial de inventario se conserva."
        )
        msg.setStyleSheet("""
            QMessageBox { background-color: #f5f5f5; }
            QLabel { color: #2c2c2c; font-size: 13px; min-width: 320px; }
            QPushButton {
                background-color: #2c3e50; color: white;
                padding: 8px 20px; border: none; border-radius: 4px;
                font-weight: bold; font-size: 13px; min-width: 80px;
            }
            QPushButton:hover { background-color: #3d5166; }
        """)
        btn_si = msg.addButton("Sí, eliminar", QMessageBox.YesRole)
        msg.addButton("Cancelar", QMessageBox.NoRole)
        msg.exec()
        if msg.clickedButton() != btn_si:
            return
        auditoria_service.limpiar_historial_global()
        self._aplicar_filtros_historial()

    # -----------------------------------------------------------------------
    # Pestaña Resumen de Actividad
    # -----------------------------------------------------------------------

    def _crear_tab_resumen(self):
        """Crea la pestaña de resumen de actividad por usuario."""
        tab = QWidget()
        tab.setStyleSheet("background-color: #f5f7fa;")
        outer = QVBoxLayout(tab)
        outer.setContentsMargins(24, 24, 24, 16)
        outer.setSpacing(16)

        # ── Cabecera ───────────────────────────────────────────────────────
        lbl_t = QLabel("📊  Resumen de Actividad")
        lbl_t.setFont(QFont("Arial", 17, QFont.Bold))
        lbl_t.setStyleSheet("color: #1a1a1a; background: transparent; border: none;")
        outer.addWidget(lbl_t)

        lbl_sub = QLabel("Actividad consolidada por usuario y módulo.")
        lbl_sub.setStyleSheet(
            "color: #888888; font-size: 13px; background: transparent; border: none;"
        )
        outer.addWidget(lbl_sub)

        # ── Filtros ────────────────────────────────────────────────────────
        card_f = QFrame()
        card_f.setStyleSheet(
            "QFrame { background-color: #ffffff; border: 1px solid #e4e8ef; border-radius: 10px; }"
        )
        f_outer = QHBoxLayout(card_f)
        f_outer.setContentsMargins(18, 12, 18, 12)
        f_outer.setSpacing(10)

        _cmb_ss = """
            QComboBox {
                padding: 7px 10px; border: 1px solid #d0d0d0;
                border-radius: 6px; background-color: #f9f9f9;
                color: #1a1a1a; font-size: 13px; min-width: 130px;
            }
            QComboBox:focus { border: 1px solid #2c3e50; }
            QComboBox::drop-down { border: none; }
            QComboBox QAbstractItemView {
                background-color: #ffffff; color: #1a1a1a;
                selection-background-color: #2c3e50; selection-color: white;
                border: 1px solid #d0d0d0;
            }
        """
        _inp_ss = """
            QDateEdit {
                padding: 7px 10px; border: 1px solid #d0d0d0;
                border-radius: 6px; background-color: #f9f9f9;
                color: #1a1a1a; font-size: 13px; min-width: 120px;
            }
            QDateEdit:focus { border: 1px solid #2c3e50; }
        """
        _cal_ss = """
            QCalendarWidget QAbstractItemView {
                selection-background-color: #5e88b4;
                selection-color: white;
                color: #2c2c2c;
                background-color: #eaf0f9;
            }
            QCalendarWidget QWidget {
                color: #2c2c2c;
                background-color: #eaf0f9;
            }
            QCalendarWidget QWidget#qt_calendar_navigationbar {
                background-color: #dce7f3;
            }
            QCalendarWidget QToolButton {
                color: #2c2c2c;
                background-color: #dce7f3;
            }
            QCalendarWidget QMenu {
                color: #2c2c2c;
                background-color: #f5f5f5;
            }
            QCalendarWidget QSpinBox {
                color: #2c2c2c;
                background-color: #f0f0f0;
            }
            QCalendarWidget QTableView {
                color: #2c2c2c;
            }
        """

        def _lbl_f(text):
            l = QLabel(text)
            l.setStyleSheet(
                "color: #555; font-weight: bold; font-size: 13px;"
                " background: transparent; border: none;"
            )
            return l

        f_outer.addWidget(_lbl_f("Módulo:"))
        self._resumen_cmb_modulo = QComboBox()
        self._resumen_cmb_modulo.addItems(
            ["Todos", "Clientes", "Pagos", "Membresías", "Inventario"]
        )
        self._resumen_cmb_modulo.setStyleSheet(_cmb_ss)
        f_outer.addWidget(self._resumen_cmb_modulo)

        _hoy    = QDate.currentDate()
        _primer = QDate(_hoy.year(), _hoy.month(), 1)

        f_outer.addWidget(_lbl_f("Desde:"))
        self._resumen_desde = QDateEdit()
        self._resumen_desde.setCalendarPopup(True)
        self._resumen_desde.setDisplayFormat("dd/MM/yyyy")
        self._resumen_desde.setMinimumDate(QDate(2000, 1, 1))
        self._resumen_desde.setDate(_primer)
        self._resumen_desde.setStyleSheet(_inp_ss)
        self._resumen_desde.calendarWidget().setStyleSheet(_cal_ss)
        f_outer.addWidget(self._resumen_desde)

        f_outer.addWidget(_lbl_f("Hasta:"))
        self._resumen_hasta = QDateEdit()
        self._resumen_hasta.setCalendarPopup(True)
        self._resumen_hasta.setDisplayFormat("dd/MM/yyyy")
        self._resumen_hasta.setMinimumDate(QDate(2000, 1, 1))
        self._resumen_hasta.setDate(_hoy)
        self._resumen_hasta.setStyleSheet(_inp_ss)
        self._resumen_hasta.calendarWidget().setStyleSheet(_cal_ss)
        f_outer.addWidget(self._resumen_hasta)

        btn_act = QPushButton("🔄  Actualizar")
        btn_act.setCursor(Qt.PointingHandCursor)
        btn_act.setStyleSheet("""
            QPushButton {
                background-color: #2c3e50; color: white;
                padding: 8px 18px; border: none; border-radius: 6px;
                font-weight: bold; font-size: 13px;
            }
            QPushButton:hover { background-color: #3d5166; }
        """)
        btn_act.clicked.connect(self._actualizar_resumen)
        f_outer.addWidget(btn_act)
        f_outer.addStretch()
        outer.addWidget(card_f)

        # ── KPI cards ──────────────────────────────────────────────────────
        kpi_row = QHBoxLayout()
        kpi_row.setSpacing(16)

        def _kpi_card(icon, title, attr_name, border_color):
            frame = QFrame()
            frame.setStyleSheet(f"""
                QFrame {{
                    background-color: #ffffff;
                    border: 1px solid #e2e8f0;
                    border-left: 5px solid {border_color};
                    border-radius: 10px;
                }}
            """)
            lay = QVBoxLayout(frame)
            lay.setContentsMargins(16, 14, 16, 14)
            lay.setSpacing(4)
            rt = QHBoxLayout()
            ico_lbl = QLabel(icon)
            ico_lbl.setFont(QFont("Arial", 16))
            ico_lbl.setStyleSheet("background: transparent; border: none;")
            t_lbl = QLabel(title)
            t_lbl.setStyleSheet(
                "color: #64748b; font-size: 11px; font-weight: bold;"
                " background: transparent; border: none;"
            )
            rt.addWidget(ico_lbl)
            rt.addWidget(t_lbl)
            rt.addStretch()
            lay.addLayout(rt)
            val = QLabel("—")
            val.setFont(QFont("Arial", 17, QFont.Bold))
            val.setStyleSheet("color: #1a2332; background: transparent; border: none;")
            val.setWordWrap(True)
            setattr(self, attr_name, val)
            lay.addWidget(val)
            return frame

        kpi_row.addWidget(_kpi_card("📋", "TOTAL ACCIONES",     "_kpi_total_lbl",   "#2563eb"))
        kpi_row.addWidget(_kpi_card("🏆", "USUARIO MÁS ACTIVO", "_kpi_usuario_lbl", "#f59e0b"))
        kpi_row.addWidget(_kpi_card("📦", "MÓDULO MÁS ACTIVO",  "_kpi_modulo_lbl",  "#16a34a"))
        outer.addLayout(kpi_row)

        # ── Título ranking ─────────────────────────────────────────────────
        lbl_rank_t = QLabel("👥  Ranking de empleados")
        lbl_rank_t.setFont(QFont("Arial", 14, QFont.Bold))
        lbl_rank_t.setStyleSheet(
            "color: #1a2332; background: transparent; border: none;"
        )
        outer.addWidget(lbl_rank_t)

        # ── Scroll de tarjetas ─────────────────────────────────────────────
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setStyleSheet("""
            QScrollArea { border: none; background-color: #f5f7fa; }
            QScrollBar:vertical { background: #f0f0f0; width: 8px; border-radius: 4px; }
            QScrollBar::handle:vertical { background: #c0c0c0; border-radius: 4px; min-height: 26px; }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }
        """)
        self._resumen_cards_widget = QWidget()
        self._resumen_cards_widget.setStyleSheet("background-color: #f5f7fa;")
        self._resumen_cards_layout = QVBoxLayout(self._resumen_cards_widget)
        self._resumen_cards_layout.setContentsMargins(0, 0, 8, 8)
        self._resumen_cards_layout.setSpacing(10)
        scroll.setWidget(self._resumen_cards_widget)
        outer.addWidget(scroll, 1)

        # Conectar filtros
        self._resumen_cmb_modulo.currentIndexChanged.connect(self._actualizar_resumen)
        self._resumen_desde.dateChanged.connect(self._actualizar_resumen)
        self._resumen_hasta.dateChanged.connect(self._actualizar_resumen)

        # Carga inicial
        self._actualizar_resumen()

        return tab

    def _actualizar_resumen(self):
        """Recalcula y redibuja el contenido del resumen de actividad."""
        modulo    = self._resumen_cmb_modulo.currentText()
        desde_str = self._resumen_desde.date().toString("yyyy-MM-dd")
        hasta_str = self._resumen_hasta.date().toString("yyyy-MM-dd")

        registros = auditoria_service.obtener_historial(
            modulo=modulo if modulo != "Todos" else None,
            limite=50000,
        )
        registros = [
            r for r in registros
            if desde_str <= r.get("fecha_hora", "")[:10] <= hasta_str
        ]

        user_totals  = defaultdict(int)
        user_modules = defaultdict(lambda: defaultdict(int))
        mod_totals   = defaultdict(int)
        for r in registros:
            usr = r.get("usuario", "")
            mod = r.get("modulo", "")
            if usr:
                user_totals[usr] += 1
                user_modules[usr][mod] += 1
            if mod:
                mod_totals[mod] += 1

        sorted_users   = sorted(user_totals.keys(), key=lambda u: -user_totals[u])
        all_users_map  = {u['username']: u for u in get_all_users()}

        # KPIs
        total_acc = sum(user_totals.values())
        self._kpi_total_lbl.setText(f"{total_acc:,}")

        if sorted_users:
            top_u    = sorted_users[0]
            top_name = (all_users_map.get(top_u, {}).get('full_name') or top_u)
            self._kpi_usuario_lbl.setText(f"{top_name}\n{user_totals[top_u]:,} acciones")
        else:
            self._kpi_usuario_lbl.setText("—")

        if mod_totals:
            top_mod = max(mod_totals, key=mod_totals.get)
            self._kpi_modulo_lbl.setText(f"{top_mod}\n{mod_totals[top_mod]:,} acciones")
        else:
            self._kpi_modulo_lbl.setText("—")

        # Limpiar tarjetas
        while self._resumen_cards_layout.count():
            item = self._resumen_cards_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        if not sorted_users:
            lbl_empty = QLabel("No hay registros para los filtros seleccionados.")
            lbl_empty.setAlignment(Qt.AlignCenter)
            lbl_empty.setStyleSheet(
                "color: #aaaaaa; font-size: 14px; padding: 40px;"
                " background: transparent; border: none;"
            )
            self._resumen_cards_layout.addWidget(lbl_empty)
        else:
            total_ranking = len(sorted_users)
            for rank, username in enumerate(sorted_users, 1):
                udata  = all_users_map.get(username, {
                    "username": username, "full_name": username, "role": "—"
                })
                modulos = dict(user_modules[username])
                card    = self._crear_empleado_card(
                    udata, modulos, rank, total_ranking, user_totals[username]
                )
                self._resumen_cards_layout.addWidget(card)

        self._resumen_cards_layout.addStretch()

    def _crear_empleado_card(self, udata: dict, modulos_data: dict,
                              ranking: int, total_ranking: int,
                              total_acciones: int) -> QFrame:
        """Devuelve una tarjeta clickeable con el resumen del empleado."""
        username  = udata.get('username', '')
        full_name = udata.get('full_name') or username
        role      = udata.get('role', '')

        rank_border = {1: "#f59e0b", 2: "#94a3b8", 3: "#b45309"}.get(ranking, "#e2e8f0")

        card = QFrame()
        card.setObjectName("empleadoCard")
        card.setStyleSheet(f"""
            QFrame#empleadoCard {{
                background-color: #ffffff;
                border: 1.5px solid {rank_border};
                border-radius: 10px;
            }}
            QFrame#empleadoCard:hover {{
                background-color: #f8faff;
                border: 1.5px solid #2563eb;
            }}
        """)
        card.setCursor(Qt.PointingHandCursor)

        lay = QHBoxLayout(card)
        lay.setContentsMargins(16, 14, 16, 14)
        lay.setSpacing(16)

        # Medalla de rango
        medal_txt = {1: "🥇", 2: "🥈", 3: "🥉"}.get(ranking, f"#{ranking}")
        lbl_med = QLabel(medal_txt)
        lbl_med.setFont(QFont("Arial", 18 if ranking <= 3 else 14, QFont.Bold))
        lbl_med.setFixedWidth(42)
        lbl_med.setAlignment(Qt.AlignCenter)
        lbl_med.setStyleSheet("background: transparent; border: none;")
        lay.addWidget(lbl_med)

        # Avatar con inicial
        av_color = _AVATAR_PALETTE[sum(ord(c) for c in username) % len(_AVATAR_PALETTE)]
        initials  = (full_name[:1] if full_name else "?").upper()
        avatar = QLabel(initials)
        avatar.setFixedSize(44, 44)
        avatar.setAlignment(Qt.AlignCenter)
        avatar.setStyleSheet(
            f"background-color: {av_color}; color: white; font-size: 18px;"
            " font-weight: bold; border-radius: 22px; border: none;"
        )
        lay.addWidget(avatar)

        # Nombre, usuario y chips de módulo
        info_l = QVBoxLayout()
        info_l.setSpacing(3)

        lbl_name = QLabel(full_name)
        lbl_name.setFont(QFont("Arial", 13, QFont.Bold))
        lbl_name.setStyleSheet("color: #1a2332; background: transparent; border: none;")

        lbl_meta = QLabel(f"@{username}  ·  {role}")
        lbl_meta.setStyleSheet(
            "color: #94a3b8; font-size: 12px; background: transparent; border: none;"
        )

        chips_row = QHBoxLayout()
        chips_row.setSpacing(6)
        for mod, cnt in sorted(modulos_data.items(), key=lambda x: -x[1])[:4]:
            bg, fg = _COLORES_MOD_BADGE.get(mod, ("#f1f5f9", "#475569"))
            chip = QLabel(f"{mod}: {cnt}")
            chip.setStyleSheet(
                f"background-color: {bg}; color: {fg}; font-size: 11px; font-weight: bold;"
                " padding: 2px 8px; border-radius: 10px; border: none;"
            )
            chips_row.addWidget(chip)
        chips_row.addStretch()

        info_l.addWidget(lbl_name)
        info_l.addWidget(lbl_meta)
        info_l.addLayout(chips_row)
        lay.addLayout(info_l, 1)

        # Total de acciones a la derecha
        right_l = QVBoxLayout()
        right_l.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        right_l.setSpacing(0)

        lbl_total = QLabel(f"{total_acciones:,}")
        lbl_total.setFont(QFont("Arial", 22, QFont.Bold))
        lbl_total.setAlignment(Qt.AlignRight)
        lbl_total.setStyleSheet("color: #1a2332; background: transparent; border: none;")

        lbl_acc_txt = QLabel("acciones")
        lbl_acc_txt.setAlignment(Qt.AlignRight)
        lbl_acc_txt.setStyleSheet(
            "color: #94a3b8; font-size: 11px; background: transparent; border: none;"
        )
        right_l.addStretch()
        right_l.addWidget(lbl_total)
        right_l.addWidget(lbl_acc_txt)

        if ranking == 1:
            lbl_top = QLabel("✦ TOP #1")
            lbl_top.setAlignment(Qt.AlignRight)
            lbl_top.setStyleSheet(
                "background-color: #fef3c7; color: #b45309;"
                " font-size: 11px; font-weight: bold;"
                " padding: 3px 10px; border-radius: 10px; border: none;"
            )
            right_l.addWidget(lbl_top)

        right_l.addStretch()
        lay.addLayout(right_l)

        # Conectar click para abrir perfil
        def _abrir(_event,
                   _ud=udata, _mods=modulos_data,
                   _rk=ranking, _tot=total_ranking):
            dlg = PerfilEmpleadoDialog(_ud, _mods, _rk, _tot, parent=self)
            dlg.exec()

        card.mousePressEvent = _abrir
        return card

    def cerrar_sesion(self):
        """Cierra la sesión actual y regresa al login."""
        msg = QMessageBox(self)
        msg.setWindowTitle("Confirmar")
        msg.setText("¿Desea cerrar la sesión actual?")
        msg.setStyleSheet("""
            QMessageBox { background-color: #f5f5f5; }
            QLabel { color: #2c2c2c; font-size: 14px; min-width: 280px; border: none; }
            QPushButton {
                background-color: #2c3e50;
                color: white;
                padding: 8px 20px;
                border: none;
                border-radius: 4px;
                font-weight: bold;
                font-size: 13px;
                min-width: 80px;
            }
            QPushButton:hover { background-color: #3d5166; }
        """)
        btn_si = msg.addButton("Sí", QMessageBox.YesRole)
        msg.addButton("No", QMessageBox.NoRole)
        msg.exec()
        if msg.clickedButton() != btn_si:
            return

        # Emitir señal para que MainWindow gestione el regreso al login
        self.logout_solicitado.emit()

